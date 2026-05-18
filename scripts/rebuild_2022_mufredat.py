# -*- coding: utf-8 -*-
"""
2022 müfredatını öğrenci veri setine göre yeniden kur.

Adımlar:
 1) 2022_ogrenci_not_veri_seti.xlsx'ten 9 bölüm × 2 dönem × 4 ders + istatistik oku
 2) data/2022_Mufredat.xlsx dosyasını DB Türkçe ad'larıyla yeniden yaz (18 satır)
 3) DB yedeği al
 4) 2022 dışı tüm mufredat + mufredat_ders sil; ders_kriterleri tamamen sil
 5) 72 ders için ders kaydı garanti et (yoksa Seçmeli oluştur)
 6) 2022 için mufredat (18) + mufredat_ders bağlarını kur
 7) Müfredat dersleri status=1, ilgili fakültedeki diğer Seçmeli'ler status=0 (havuz)
 8) Ders Analizi sekmesinden ders_kriterleri üret (2022)
"""
import os
import shutil
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook

KOK = Path(__file__).parent.parent
DB = KOK / "data" / "adil_secmeli.db"
OGR = KOK / "data" / "2022_ogrenci_not_veri_seti.xlsx"
MUF = KOK / "data" / "2022_Mufredat.xlsx"
YIL = 2022


def oku_veri_seti():
    """Ana Veri -> ders listesi; Ders Analizi -> istatistik."""
    wb = openpyxl.load_workbook(OGR, read_only=True)

    ws = wb["Ana Veri"]
    it = ws.iter_rows(min_row=1, values_only=True)
    hdr = list(next(it))
    ix = {h: i for i, h in enumerate(hdr)}
    # (bolum_id, donem) -> list[(kod, ad, kredi)]
    dersler = defaultdict(dict)
    for r in it:
        bid = int(r[ix["bolum_id"]])
        don = str(r[ix["donem"]]).strip()
        kod = str(r[ix["ders_kodu"]]).strip()
        ad = str(r[ix["ders_adi"]]).strip()
        kredi = int(r[ix["kredi"]] or 3)
        dersler[(bid, don)][kod] = (ad, kredi)

    ws2 = wb["Ders Analizi"]
    it2 = ws2.iter_rows(min_row=1, values_only=True)
    h2 = list(next(it2))
    j = {h: i for i, h in enumerate(h2)}
    # kod -> istatistik (kod benzersiz)
    stat = {}
    for r in it2:
        kod = str(r[j["ders_kodu"]]).strip()
        stat[kod] = {
            "bolum_id": int(r[j["bolum_id"]]),
            "donem": str(r[j["donem"]]).strip(),
            "kayit": int(r[j["kayit_sayisi"]] or 0),
            "gecme_orani": float(r[j["gecme_orani_%"]] or 0),
            "ort_agirlikli": float(r[j["ort_agirlikli"]] or 0),
            "katilim_yuzde": float(r[j["ort_katilim_yuzde"]] or 0),
        }
    wb.close()
    return dersler, stat


def bolum_fakulte_map(cur):
    cur.execute("SELECT bolum_id, fakulte_id, ad FROM bolum")
    b = {int(r[0]): (int(r[1]), r[2]) for r in cur.fetchall()}
    cur.execute("SELECT fakulte_id, ad FROM fakulte")
    f = {int(r[0]): r[1] for r in cur.fetchall()}
    return b, f


def yaz_mufredat_excel(dersler, bmap, fmap):
    """18 satır: 9 bölüm × (Guz, Bahar), 4 seçmeli ders."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sayfa1"
    ws.append([
        "ID", "Fakülte", "Bölüm", "Yıl", "Dönem",
        "Seçmeli Ders 1", "Seçmeli Ders 2", "Seçmeli Ders 3", "Seçmeli Ders 4",
    ])
    rid = 1
    for bid in sorted({b for (b, _d) in dersler}):
        fak_id, bol_ad = bmap[bid]
        fak_ad = fmap[fak_id]
        for don in ("Guz", "Bahar"):
            kayit = dersler.get((bid, don), {})
            adlar = [v[0] for v in list(kayit.values())[:4]]
            while len(adlar) < 4:
                adlar.append("")
            ws.append([rid, fak_ad, bol_ad, YIL, don, *adlar])
            rid += 1
    wb.save(str(MUF))
    print(f"  2022_Mufredat.xlsx yazildi: {rid - 1} satir")


def main():
    print("1) Veri seti okunuyor...")
    dersler, stat = oku_veri_seti()
    print(f"   {len(dersler)} (bolum,donem) kombinasyonu, {len(stat)} benzersiz ders")

    conn = sqlite3.connect(str(DB))
    cur = conn.cursor()
    bmap, fmap = bolum_fakulte_map(cur)

    print("2) 2022_Mufredat.xlsx yeniden yaziliyor...")
    yaz_mufredat_excel(dersler, bmap, fmap)

    print("3) DB yedegi aliniyor...")
    yedek = str(DB) + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(str(DB), yedek)
    print(f"   Yedek: {os.path.basename(yedek)}")

    print("4) 2022 disi mufredat + tum ders_kriterleri siliniyor...")
    cur.execute(
        "DELETE FROM mufredat_ders WHERE mufredat_id IN "
        "(SELECT mufredat_id FROM mufredat WHERE akademik_yil != ?)",
        (YIL,),
    )
    cur.execute("DELETE FROM mufredat WHERE akademik_yil != ?", (YIL,))
    # Kalan 2022 mufredat'i da sifirdan kuracagiz
    cur.execute(
        "DELETE FROM mufredat_ders WHERE mufredat_id IN "
        "(SELECT mufredat_id FROM mufredat WHERE akademik_yil = ?)",
        (YIL,),
    )
    cur.execute("DELETE FROM mufredat WHERE akademik_yil = ?", (YIL,))
    cur.execute("DELETE FROM ders_kriterleri")
    conn.commit()

    print("5) 72 ders icin ders kaydi garanti ediliyor...")
    kod_to_dersid = {}
    olusturulan = 0
    for kod, st in stat.items():
        bid = st["bolum_id"]
        fak_id, _ = bmap[bid]
        # ders_adi/kredi (Ana Veri'den)
        ad, kredi = None, 3
        for (b, d), kayit in dersler.items():
            if kod in kayit:
                ad, kredi = kayit[kod]
                break
        cur.execute("SELECT ders_id FROM ders WHERE kod = ?", (kod,))
        row = cur.fetchone()
        if row:
            did = int(row[0])
            cur.execute(
                "UPDATE ders SET ad=?, bolum_id=?, fakulte_id=?, kredi=?, "
                "DersTipi='Seçmeli', status=1 WHERE ders_id=?",
                (ad, bid, fak_id, kredi, did),
            )
        else:
            cur.execute(
                "INSERT INTO ders (bolum_id, fakulte_id, ad, kredi, kontenjan, "
                "alan, status, akts, DersTipi, kod) "
                "VALUES (?, ?, ?, ?, ?, ?, 1, ?, 'Seçmeli', ?)",
                (bid, fak_id, ad, kredi, 60, "Seçmeli", kredi, kod),
            )
            did = int(cur.lastrowid)
            olusturulan += 1
        kod_to_dersid[kod] = did
    conn.commit()
    print(f"   {olusturulan} yeni ders olusturuldu, {len(kod_to_dersid)} ders hazir")

    print("6) 2022 mufredat + mufredat_ders kuruluyor...")
    mufredat_sayisi = 0
    for bid in sorted({b for (b, _d) in dersler}):
        fak_id, _ = bmap[bid]
        for don in ("Guz", "Bahar"):
            kayit = dersler.get((bid, don), {})
            if not kayit:
                continue
            cur.execute(
                "INSERT INTO mufredat (fakulte_id, akademik_yil, bolum_id, "
                "donem, durum, versiyon) VALUES (?, ?, ?, ?, 'aktif', 1)",
                (fak_id, YIL, bid, don),
            )
            mid = int(cur.lastrowid)
            mufredat_sayisi += 1
            for kod in kayit:
                did = kod_to_dersid.get(kod)
                if did:
                    cur.execute(
                        "INSERT INTO mufredat_ders (mufredat_id, ders_id) "
                        "VALUES (?, ?)",
                        (mid, did),
                    )
    conn.commit()
    print(f"   {mufredat_sayisi} mufredat satiri olusturuldu")

    print("7) Status ayarlaniyor (mufredat=1, havuz=0)...")
    mufredat_dersleri = set(kod_to_dersid.values())
    ilgili_fakulteler = {bmap[b][0] for (b, _d) in dersler}
    # Once ilgili fakultelerdeki tum Secmeli dersleri havuza (0)
    fak_list = ",".join(str(f) for f in ilgili_fakulteler)
    cur.execute(
        f"UPDATE ders SET status=0 WHERE DersTipi='Seçmeli' "
        f"AND fakulte_id IN ({fak_list})"
    )
    # Mufredattakileri 1 yap
    if mufredat_dersleri:
        q = ",".join("?" * len(mufredat_dersleri))
        cur.execute(
            f"UPDATE ders SET status=1 WHERE ders_id IN ({q})",
            tuple(mufredat_dersleri),
        )
    conn.commit()
    cur.execute(
        f"SELECT status, COUNT(*) FROM ders WHERE DersTipi='Seçmeli' "
        f"AND fakulte_id IN ({fak_list}) GROUP BY status"
    )
    print(f"   Secmeli ders status dagilimi: {cur.fetchall()}")

    print("8) ders_kriterleri uretiliyor (Ders Analizi'nden)...")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    krit_sayisi = 0
    for kod, st in stat.items():
        did = kod_to_dersid.get(kod)
        if not did:
            continue
        kayit = st["kayit"]
        gecen = round(kayit * st["gecme_orani"] / 100.0)
        basari = round(st["ort_agirlikli"], 2)
        anket_kat = round(kayit * st["katilim_yuzde"] / 100.0)
        cur.execute(
            "INSERT INTO ders_kriterleri "
            "(ders_id, yil, donem, toplam_ogrenci, gecen_ogrenci, "
            "basari_ortalamasi, kontenjan, kayitli_ogrenci, anket_katilimci, "
            "anket_dersi_secen, anket_veri_kaynagi, criteria_veri_kaynagi, "
            "criteria_updated_at, is_active) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, "
            "'ogrenci_veri_seti_2022', 'ogrenci_veri_seti_2022', ?, 1)",
            (did, YIL, st["donem"], kayit, gecen, basari, 60, kayit,
             anket_kat, kayit, now),
        )
        krit_sayisi += 1
    conn.commit()
    print(f"   {krit_sayisi} ders_kriterleri satiri olusturuldu")

    # Ozet dogrulama
    print("\n=== DOGRULAMA ===")
    cur.execute("SELECT akademik_yil, COUNT(*) FROM mufredat GROUP BY akademik_yil")
    print("mufredat yil dagilimi:", cur.fetchall())
    cur.execute("SELECT DISTINCT donem FROM mufredat")
    print("mufredat donem:", [r[0] for r in cur.fetchall()])
    cur.execute("SELECT COUNT(*) FROM mufredat_ders")
    print("mufredat_ders toplam:", cur.fetchone()[0])
    cur.execute("SELECT yil, COUNT(*) FROM ders_kriterleri GROUP BY yil")
    print("ders_kriterleri yil:", cur.fetchall())
    conn.close()
    print("\nTAMAMLANDI.")


if __name__ == "__main__":
    main()
