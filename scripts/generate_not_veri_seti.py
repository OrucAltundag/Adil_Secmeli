# -*- coding: utf-8 -*-
"""
Ogrenci not veri seti uretici (yil-parametreli)
9 bolum x 50 ogrenci x 8 ders = 3600 satir
Cikti: data/<yil>_ogrenci_not_veri_seti.xlsx  (4 sekme)

Kullanim:
    python -m scripts.generate_not_veri_seti                # 2022 (varsayilan)
    python -m scripts.generate_not_veri_seti --yil 2021
    python -m scripts.generate_not_veri_seti --yil 2023
    python -m scripts.generate_not_veri_seti --tum-yillar   # 2021, 2022, 2023

Not: 2022 uretimi seed=42 ile birebir korunur (mevcut dosyayla ayni).
Diger yillarda seed=yil ve bolum bazli trend kaymasi uygulanir; boylece
trend analizi algoritmasi gercek cok-yilli sinyalle calisir.
"""
import argparse
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Trend referans yili — bu yilda kayma 0'dir.
TREND_REFERANS_YIL = 2022

# ─── Sabitler ────────────────────────────────────────────────────────────────

ERKEK_ADLAR = [
    "Ahmet", "Mehmet", "Ali", "Mustafa", "Ibrahim", "Huseyin", "Hasan",
    "Yusuf", "Emre", "Furkan", "Berkay", "Onur", "Can", "Burak", "Mert",
    "Alp", "Kerem", "Serkan", "Berk", "Omer",
]
KADIN_ADLAR = [
    "Ayse", "Fatma", "Zeynep", "Elif", "Merve", "Selin", "Irem", "Busra",
    "Ozlem", "Esra", "Buse", "Ece", "Dilan", "Cansu", "Melis",
    "Yaren", "Nisa", "Hilal", "Pinar", "Beyza",
]
SOYADLAR = [
    "Yilmaz", "Kaya", "Demir", "Celik", "Sahin", "Yildiz", "Ozturk",
    "Aydin", "Arslan", "Dogan", "Kilic", "Aslan", "Cetin", "Koc",
    "Kurt", "Oz", "Acar", "Bulut", "Polat", "Erdogan", "Gunduz",
    "Kaplan", "Karaca", "Guler", "Akgun",
]

BOLUMLER = {
    1: {
        "adi": "Tip Fakultesi", "fakulte": "Tip Fakultesi", "zorluk": 1.15,
        "dersler_guz": [
            ("TIP101S", "Klinik Anatomi Secmeli", 3),
            ("TIP102S", "Tibbi Etik Secmeli", 2),
            ("OZD061",  "Girisimcilik", 2),
            ("OZD063",  "Toplum Projesi", 2),
        ],
        "dersler_bahar": [
            ("TIP609",  "Klinik Farmakoloji", 3),
            ("TIP610",  "Ileri Fizyoloji", 3),
            ("OZD062",  "Iletisim Becerileri", 2),
            ("OZD064",  "Arastirma Yontemleri", 2),
        ],
    },
    2: {
        "adi": "Bilgisayar Muhendisligi", "fakulte": "Muhendislik Fakultesi", "zorluk": 1.10,
        "dersler_guz": [
            ("BMB401S", "Makine Ogrenmesi", 3),
            ("BMB403S", "Bulut Bilisim", 3),
            ("BMB405S", "Mobil Uygulama Gelistirme", 3),
            ("BMB407S", "Siber Guvenlik", 3),
        ],
        "dersler_bahar": [
            ("BMB402S", "Derin Ogrenme", 3),
            ("BMB404S", "Buyuk Veri Analizi", 3),
            ("BMB406S", "Oyun Gelistirme", 3),
            ("BMB408S", "Dogal Dil Isleme", 3),
        ],
    },
    3: {
        "adi": "Elektrik-Elektronik Muhendisligi", "fakulte": "Muhendislik Fakultesi", "zorluk": 1.12,
        "dersler_guz": [
            ("EEM401S", "Kablosuz Haberlesme", 3),
            ("EEM403S", "Guc Elektroniği", 3),
            ("EEM405S", "Goruntu Isleme", 3),
            ("EEM407S", "Gomulu Sistemler", 3),
        ],
        "dersler_bahar": [
            ("EEM402S", "Nesnelerin Interneti", 3),
            ("EEM404S", "Yenilenebilir Enerji", 3),
            ("EEM406S", "Robot Kontrolu", 3),
            ("EEM408S", "Sinyal Isleme", 3),
        ],
    },
    4: {
        "adi": "Endustri Muhendisligi", "fakulte": "Muhendislik Fakultesi", "zorluk": 1.05,
        "dersler_guz": [
            ("END401S", "Tedarik Zinciri Yonetimi", 3),
            ("END403S", "Proje Yonetimi", 3),
            ("END405S", "Sezgisel Optimizasyon", 3),
            ("END407S", "Kalite Yonetim Sistemleri", 3),
        ],
        "dersler_bahar": [
            ("END402S", "Yapay Zeka Uygulamalari", 3),
            ("END404S", "Veri Madenciligi", 3),
            ("END406S", "Ileri Uretim Sistemleri", 3),
            ("END408S", "Surec Simulasyonu", 3),
        ],
    },
    5: {
        "adi": "Hemsirelik", "fakulte": "Saglik Bilimleri Fakultesi", "zorluk": 0.98,
        "dersler_guz": [
            ("HEM401S", "Ileri Klinik Hemsirelik", 3),
            ("HEM403S", "Psikiyatri Hemsireliginde Guncel", 3),
            ("HEM405S", "Onkoloji Hemsireliginde Kanit", 3),
            ("HEM407S", "Aile Sagligi Hemsireliginde Guncel", 3),
        ],
        "dersler_bahar": [
            ("HEM402S", "Kadin Sagligi Hemsireliginde Guncel", 3),
            ("HEM404S", "Cocuk Sagligi Hemsireliginde Guncel", 3),
            ("HEM406S", "Ameliyathane Hemsireliginde Guncel", 3),
            ("HEM408S", "Yogun Bakim Hemsireliginde Guncel", 3),
        ],
    },
    6: {
        "adi": "Ebelik", "fakulte": "Saglik Bilimleri Fakultesi", "zorluk": 0.96,
        "dersler_guz": [
            ("EBE401S", "Yuksek Riskli Gebelik Bakimi", 3),
            ("EBE403S", "Dogum Agi Yonetimi", 3),
            ("EBE405S", "Neonatal Bakim", 3),
            ("EBE407S", "Reproductive Saglik", 3),
        ],
        "dersler_bahar": [
            ("EBE402S", "Infertilite Hemsireliginde Guncel", 3),
            ("EBE404S", "Maternal Saglik", 3),
            ("EBE406S", "Postpartum Bakim", 3),
            ("EBE408S", "Kadinlarda Onkoloji", 3),
        ],
    },
    7: {
        "adi": "Fizyoterapi ve Rehabilitasyon", "fakulte": "Saglik Bilimleri Fakultesi", "zorluk": 1.00,
        "dersler_guz": [
            ("FTR608S", "Kardiyopulmoner Rehabilitasyon", 3),
            ("FTR609S", "Pediatrik Rehabilitasyon", 3),
            ("FTR611S", "Manuel Terapi Teknikleri", 3),
            ("FTR613S", "Sportif Rehabilitasyon", 3),
        ],
        "dersler_bahar": [
            ("FTR602S", "Norolojik Rehabilitasyon Guncel", 3),
            ("FTR604S", "Ortopedik Rehabilitasyon Guncel", 3),
            ("FTR606S", "Onkolojik Rehabilitasyon", 3),
            ("FTR610S", "Geriatrik Rehabilitasyon", 3),
        ],
    },
    8: {
        "adi": "Gastronomi ve Mutfak Sanatlari", "fakulte": "Turizm Fakultesi", "zorluk": 0.88,
        "dersler_guz": [
            ("GAS401S", "Molekuler Gastronomi", 3),
            ("GAS403S", "Dunya Mutfaklari", 3),
            ("GAS405S", "Seker ve Cikolata Sanatlari", 3),
            ("GAS407S", "Restoran Yonetimi", 3),
        ],
        "dersler_bahar": [
            ("GAS402S", "Fermentasyon ve Mutfak Kimyasi", 3),
            ("GAS404S", "Pastane ve Firinci Sanatlari", 3),
            ("GAS406S", "Gida Stili ve Fotografciligi", 3),
            ("GAS408S", "Gastronomi Turizmi", 3),
        ],
    },
    9: {
        "adi": "Ilahiyat", "fakulte": "Ilahiyat Fakultesi", "zorluk": 0.92,
        "dersler_guz": [
            ("ILH109S", "Islami Etik ve Ahlak", 3),
            ("ILH110S", "Karsilastirmali Dinler Tarihi", 3),
            ("ILH111S", "Hadis Ilimleri Secmeli", 3),
            ("ILH112S", "Kelam Tarihi Secmeli", 3),
        ],
        "dersler_bahar": [
            ("ILH201S", "Kuran Yorumlari ve Tefsir", 3),
            ("ILH202S", "Islam Felsefesi", 3),
            ("ILH203S", "Din Egitimi Yontemleri", 3),
            ("ILH204S", "Tasavvuf ve Manevi Hayat", 3),
        ],
    },
}

HARF_ESIK = [
    (90, "AA", 4.0), (85, "BA", 3.5), (80, "BB", 3.0),
    (75, "CB", 2.5), (70, "CC", 2.0), (65, "DC", 1.5),
    (60, "DD", 1.0), (0,  "FF", 0.0),
]

ANA_VERI_SUTUNLAR = [
    "ogrenci_no", "ad", "soyad", "cinsiyet", "dogum_yili", "sinif",
    "burslu_mu", "bolum_id", "bolum_adi", "fakulte_adi",
    "akademik_yil", "donem", "ders_id", "ders_kodu", "ders_adi", "kredi",
    "vize_notu", "proje_notu", "final_notu", "agirlikli_not",
    "gecme_esigi", "harf_notu", "gano_katkisi_4luk",
    "katilim_sayisi", "toplam_hafta", "katilim_yuzdesi",
    "devamsiz_mi", "gecti_mi", "basari_durumu",
    "begen_puani_1_5", "zorluk_alg_1_5", "kariyer_katkisi_1_5",
    "ilgi_alani_uyumu_1_5", "tekrar_alacak_mi", "yeniden_alir_mi",
    "not_tutarsizlik_flag",
]


# ─── Yardimci fonksiyonlar ───────────────────────────────────────────────────

def clamp(v: float, lo: float = 0.0, hi: float = 100.0) -> float:
    return max(lo, min(hi, v))


def gauss_clamp(mean: float, std: float, lo: float = 0.0, hi: float = 100.0) -> float:
    return round(clamp(random.gauss(mean, std), lo, hi), 1)


def bolum_yil_drift(bolum_id: int, yil: int) -> float:
    """Bolum bazli yillik performans kaymasi (trend sinyali).

    Referans yil (2022) icin 0 doner. Bazi bolumler yukari (yukselen ders
    talebi/basari), bazilari asagi trend ile ayrisir; boylece trend analizi
    algoritmasi farkli etiketler (yukselen/dusen/sabit) uretebilir.
    """
    yil_farki = yil - TREND_REFERANS_YIL
    if yil_farki == 0:
        return 0.0
    # Bolum id'ye gore yon: tek id'ler yukselen, cift id'ler dusen egilim.
    yon = 1.0 if (bolum_id % 2 == 1) else -1.0
    # Bolum bazli buyukluk (1.0–2.2 puan/yil), deterministik.
    buyukluk = 1.0 + ((bolum_id * 7) % 13) / 10.0
    return round(yon * buyukluk * yil_farki, 2)


def harf_not(agirlikli):
    for esik, harf, gano in HARF_ESIK:
        if agirlikli >= esik:
            return harf, gano
    return "FF", 0.0


def uret_satir(
    ogrenci_no, ad, soyad, cinsiyet, dogum_yili, sinif, burslu,
    bolum_id, bolum_adi, fakulte_adi,
    donem, ders_id, ders_kodu, ders_adi, kredi,
    perf, bolum_zorluk, ders_varyasyon,
    akademik_yil="2022-2023",
):
    eff = clamp(perf * bolum_zorluk * ders_varyasyon, 30, 97)

    vize   = gauss_clamp(eff * 0.95, 12)
    proje  = gauss_clamp(eff * 1.05, 10)
    final  = gauss_clamp(eff * 1.00, 13)
    agirlik = round(0.40 * vize + 0.10 * proje + 0.50 * final, 2)
    gecme_esigi = 60.0
    harf, gano4 = harf_not(agirlik)

    toplam_hafta = 14
    if agirlik >= 70:
        kat_ort, kat_std = 12.5, 1.2
    elif agirlik >= 50:
        kat_ort, kat_std = 11.0, 1.8
    else:
        kat_ort, kat_std = 9.5, 2.5
    katilim = int(clamp(round(random.gauss(kat_ort, kat_std)), 0, toplam_hafta))
    katilim_yuzde = round(katilim / toplam_hafta * 100, 1)
    devamsiz = katilim < 10

    if devamsiz:
        harf, gano4 = "FF", 0.0
        gecti = False
    else:
        gecti = agirlik >= gecme_esigi

    begen     = max(1, min(5, round(random.gauss(1 + agirlik / 28, 0.8))))
    zorluk_a  = max(1, min(5, round(random.gauss(5 - agirlik / 28, 0.9))))
    kariyer   = max(1, min(5, round(random.gauss(3.2, 0.8))))
    ilgi      = max(1, min(5, round(random.gauss(3.0, 0.9))))
    tekrar    = "Evet" if begen >= 4 else ("Belki" if begen == 3 else "Hayir")
    yeniden   = "Evet" if (gecti and begen >= 3) else "Hayir"
    tutarsiz  = (agirlik >= 75 and katilim <= 8) or (agirlik < 45 and katilim >= 13)

    return {
        "ogrenci_no": ogrenci_no, "ad": ad, "soyad": soyad,
        "cinsiyet": cinsiyet, "dogum_yili": dogum_yili, "sinif": sinif,
        "burslu_mu": "Evet" if burslu else "Hayir",
        "bolum_id": bolum_id, "bolum_adi": bolum_adi, "fakulte_adi": fakulte_adi,
        "akademik_yil": akademik_yil, "donem": donem,
        "ders_id": ders_id, "ders_kodu": ders_kodu, "ders_adi": ders_adi, "kredi": kredi,
        "vize_notu": vize, "proje_notu": proje, "final_notu": final,
        "agirlikli_not": agirlik, "gecme_esigi": gecme_esigi,
        "harf_notu": harf, "gano_katkisi_4luk": gano4,
        "katilim_sayisi": katilim, "toplam_hafta": toplam_hafta,
        "katilim_yuzdesi": katilim_yuzde,
        "devamsiz_mi": "Evet" if devamsiz else "Hayir",
        "gecti_mi": "Evet" if gecti else "Hayir",
        "basari_durumu": "Gecti" if gecti else "Kaldi",
        "begen_puani_1_5": begen, "zorluk_alg_1_5": zorluk_a,
        "kariyer_katkisi_1_5": kariyer, "ilgi_alani_uyumu_1_5": ilgi,
        "tekrar_alacak_mi": tekrar, "yeniden_alir_mi": yeniden,
        "not_tutarsizlik_flag": "Evet" if tutarsiz else "Hayir",
    }


def ana_veri_uret(yil: int = 2022):
    akademik_yil = f"{yil}-{yil + 1}"
    rows = []
    counter = 1
    for bolum_id, b in BOLUMLER.items():
        drift = bolum_yil_drift(bolum_id, yil)
        tum_dersler = []
        for idx, (kod, ad, kr) in enumerate(b["dersler_guz"]):
            tum_dersler.append(("Guz",   f"D{bolum_id:02d}{idx+1:02d}", kod, ad, kr))
        for idx, (kod, ad, kr) in enumerate(b["dersler_bahar"]):
            tum_dersler.append(("Bahar", f"D{bolum_id:02d}{idx+5:02d}", kod, ad, kr))

        for _ in range(50):
            cins  = "Erkek" if random.random() < 0.50 else "Kadin"
            ad    = random.choice(ERKEK_ADLAR if cins == "Erkek" else KADIN_ADLAR)
            soyad = random.choice(SOYADLAR)
            dogum = random.randint(1998, 2004)
            sinif = random.choice([3, 4])
            burslu = random.random() < 0.25
            ono   = f"20{bolum_id:02d}{counter:04d}"
            counter += 1

            perf_base = clamp(random.gauss(72, 10) + (5 if burslu else 0) + drift, 42, 96)

            for donem, ders_id, ders_kodu, ders_adi, kredi in tum_dersler:
                varyasyon = clamp(random.gauss(1.0, 0.08), 0.82, 1.18)
                perf = clamp(perf_base + random.gauss(0, 5), 30, 97)
                satir = uret_satir(
                    ono, ad, soyad, cins, dogum, sinif, burslu,
                    bolum_id, b["adi"], b["fakulte"],
                    donem, ders_id, ders_kodu, ders_adi, kredi,
                    perf, b["zorluk"], varyasyon,
                    akademik_yil=akademik_yil,
                )
                rows.append(satir)
    return rows


# ─── Stil yardimcilari ───────────────────────────────────────────────────────

def header_style(cell, bg="1565C0"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def style_row(row_cells, even, alt_color="EBF5FB"):
    for c in row_cells:
        c.border = thin_border()
        if even:
            c.fill = PatternFill("solid", start_color=alt_color)
        c.alignment = Alignment(vertical="center")

def autofit(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(min_w, max_len + 2), max_w
        )


# ─── Sekme 1: Ana Veri ───────────────────────────────────────────────────────

def yaz_ana_veri(wb, rows):
    ws = wb.create_sheet("Ana Veri")
    for ci, col in enumerate(ANA_VERI_SUTUNLAR, 1):
        c = ws.cell(row=1, column=ci, value=col)
        header_style(c)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    GECMEDI_RENK = PatternFill("solid", start_color="FFCDD2")
    DEVAMSIZ_RENK = PatternFill("solid", start_color="FFF9C4")

    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(ANA_VERI_SUTUNLAR, 1):
            c = ws.cell(row=ri, column=ci, value=row[key])
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="EBF5FB")
        if row["devamsiz_mi"] == "Evet":
            for ci in range(1, len(ANA_VERI_SUTUNLAR) + 1):
                ws.cell(row=ri, column=ci).fill = DEVAMSIZ_RENK
        elif row["gecti_mi"] == "Hayir":
            for ci in range(1, len(ANA_VERI_SUTUNLAR) + 1):
                ws.cell(row=ri, column=ci).fill = GECMEDI_RENK

    autofit(ws)


# ─── Sekme 2: Bolum Ozeti ────────────────────────────────────────────────────

def yaz_bolum_ozeti(wb, rows):
    ws = wb.create_sheet("Bolum Ozeti")
    headers = [
        "bolum_id", "bolum_adi", "fakulte_adi",
        "ogrenci_sayisi", "ders_sayisi", "toplam_kayit",
        "ort_agirlikli_not", "ort_vize", "ort_final",
        "gecme_orani_%", "devamsizlik_orani_%", "tutarsizlik_orani_%",
        "ort_begen", "ort_zorluk_alg", "ort_kariyer",
    ]
    for ci, h in enumerate(headers, 1):
        header_style(ws.cell(row=1, column=ci, value=h), bg="1A5276")

    # bolum bazli ozet hesapla
    from collections import defaultdict
    b_data = defaultdict(list)
    for r in rows:
        b_data[r["bolum_id"]].append(r)

    for ri, (bid, grp) in enumerate(sorted(b_data.items()), 2):
        n = len(grp)
        gecen = sum(1 for r in grp if r["gecti_mi"] == "Evet")
        devamsiz = sum(1 for r in grp if r["devamsiz_mi"] == "Evet")
        tutarsiz = sum(1 for r in grp if r["not_tutarsizlik_flag"] == "Evet")
        ders_sayisi = len({r["ders_kodu"] for r in grp})
        ogrenci_sayisi = len({r["ogrenci_no"] for r in grp})

        vals = [
            bid,
            grp[0]["bolum_adi"],
            grp[0]["fakulte_adi"],
            ogrenci_sayisi,
            ders_sayisi,
            n,
            round(sum(r["agirlikli_not"] for r in grp) / n, 2),
            round(sum(r["vize_notu"] for r in grp) / n, 2),
            round(sum(r["final_notu"] for r in grp) / n, 2),
            round(gecen / n * 100, 1),
            round(devamsiz / n * 100, 1),
            round(tutarsiz / n * 100, 1),
            round(sum(r["begen_puani_1_5"] for r in grp) / n, 2),
            round(sum(r["zorluk_alg_1_5"] for r in grp) / n, 2),
            round(sum(r["kariyer_katkisi_1_5"] for r in grp) / n, 2),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = thin_border()
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="EAF2FF")

    autofit(ws)


# ─── Sekme 3: Ders Analizi ───────────────────────────────────────────────────

def yaz_ders_analizi(wb, rows):
    ws = wb.create_sheet("Ders Analizi")
    headers = [
        "bolum_id", "bolum_adi", "donem", "ders_id", "ders_kodu", "ders_adi", "kredi",
        "kayit_sayisi", "gecme_orani_%", "devamsizlik_orani_%",
        "ort_vize", "ort_proje", "ort_final", "ort_agirlikli",
        "ort_katilim_yuzde", "ort_begen", "ort_zorluk_alg", "ort_kariyer",
    ]
    for ci, h in enumerate(headers, 1):
        header_style(ws.cell(row=1, column=ci, value=h), bg="145A32")

    from collections import defaultdict
    d_data = defaultdict(list)
    for r in rows:
        d_data[(r["bolum_id"], r["donem"], r["ders_kodu"])].append(r)

    for ri, ((bid, donem, dkod), grp) in enumerate(sorted(d_data.items()), 2):
        n = len(grp)
        gecen = sum(1 for r in grp if r["gecti_mi"] == "Evet")
        devamsiz = sum(1 for r in grp if r["devamsiz_mi"] == "Evet")
        vals = [
            bid, grp[0]["bolum_adi"], donem,
            grp[0]["ders_id"], dkod, grp[0]["ders_adi"], grp[0]["kredi"],
            n,
            round(gecen / n * 100, 1),
            round(devamsiz / n * 100, 1),
            round(sum(r["vize_notu"] for r in grp) / n, 2),
            round(sum(r["proje_notu"] for r in grp) / n, 2),
            round(sum(r["final_notu"] for r in grp) / n, 2),
            round(sum(r["agirlikli_not"] for r in grp) / n, 2),
            round(sum(r["katilim_yuzdesi"] for r in grp) / n, 1),
            round(sum(r["begen_puani_1_5"] for r in grp) / n, 2),
            round(sum(r["zorluk_alg_1_5"] for r in grp) / n, 2),
            round(sum(r["kariyer_katkisi_1_5"] for r in grp) / n, 2),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = thin_border()
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="EAFAF1")

    autofit(ws)


# ─── Sekme 4: Ogrenci Ozeti ──────────────────────────────────────────────────

def yaz_ogrenci_ozeti(wb, rows):
    ws = wb.create_sheet("Ogrenci Ozeti")
    headers = [
        "ogrenci_no", "ad", "soyad", "cinsiyet", "dogum_yili", "sinif",
        "burslu_mu", "bolum_id", "bolum_adi", "fakulte_adi",
        "alinan_ders_sayisi", "gecilen_ders_sayisi", "gecme_orani_%",
        "ort_agirlikli_not", "ort_vize", "ort_final",
        "ort_katilim_yuzde", "devamsiz_ders_sayisi",
    ]
    for ci, h in enumerate(headers, 1):
        header_style(ws.cell(row=1, column=ci, value=h), bg="6C3483")

    from collections import defaultdict
    o_data = defaultdict(list)
    for r in rows:
        o_data[r["ogrenci_no"]].append(r)

    for ri, (ono, grp) in enumerate(sorted(o_data.items()), 2):
        n = len(grp)
        gecen = sum(1 for r in grp if r["gecti_mi"] == "Evet")
        devamsiz = sum(1 for r in grp if r["devamsiz_mi"] == "Evet")
        g = grp[0]
        vals = [
            ono, g["ad"], g["soyad"], g["cinsiyet"], g["dogum_yili"], g["sinif"],
            g["burslu_mu"], g["bolum_id"], g["bolum_adi"], g["fakulte_adi"],
            n, gecen,
            round(gecen / n * 100, 1),
            round(sum(r["agirlikli_not"] for r in grp) / n, 2),
            round(sum(r["vize_notu"] for r in grp) / n, 2),
            round(sum(r["final_notu"] for r in grp) / n, 2),
            round(sum(r["katilim_yuzdesi"] for r in grp) / n, 1),
            devamsiz,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = thin_border()
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="F5EEF8")

    autofit(ws)


# ─── Ana ─────────────────────────────────────────────────────────────────────

def uret_yil(yil: int) -> Path:
    """Tek bir yil icin veri setini uretir ve dosya yolunu doner."""
    # 2022 birebir korunsun diye orijinal seed (42); diger yillar seed=yil.
    random.seed(42 if yil == 2022 else yil)

    proje_kok = Path(__file__).parent.parent
    cikti = proje_kok / "data" / f"{yil}_ogrenci_not_veri_seti.xlsx"
    cikti.parent.mkdir(parents=True, exist_ok=True)

    print(f"[{yil}] Veri uretiliyor...")
    rows = ana_veri_uret(yil)
    print(f"  {len(rows)} satir uretildi.")

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)  # bos sekmeyi sil

    yaz_ana_veri(wb, rows)
    yaz_bolum_ozeti(wb, rows)
    yaz_ders_analizi(wb, rows)
    yaz_ogrenci_ozeti(wb, rows)

    wb.save(str(cikti))
    print(f"  Kaydedildi: {cikti}  (sekmeler: {wb.sheetnames})")
    return cikti


def main():
    parser = argparse.ArgumentParser(description="Ogrenci not veri seti uretici")
    parser.add_argument("--yil", type=int, default=2022)
    parser.add_argument(
        "--tum-yillar",
        action="store_true",
        help="2021, 2022 ve 2023 veri setlerini birlikte uretir.",
    )
    args = parser.parse_args()

    yillar = [2021, 2022, 2023] if args.tum_yillar else [args.yil]
    for yil in yillar:
        uret_yil(yil)


if __name__ == "__main__":
    main()
