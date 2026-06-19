# -*- coding: utf-8 -*-
"""
Anket / tercih veri seti uretici (sistemin import formatina uygun).

Senaryo (kullanici istegi):
- 2022 yilindaki dersler icin anket verisi.
- Her FAKULTE icin 100 ogrencinin anket doldurdugu varsayilir (toplam_katilimci).
- Butun dersler secilmek zorunda DEGIL: her fakulte icin 10–20 arasi ders secilir.
- Secilen dersler ogrenciler tarafindan RASTGELE oranlarda tercih edilmis kabul edilir.

Cikti: data/2022_anket_tercih_veri_seti.xlsx
  - "AnketSonuclari" sayfasi: fakulte_adi, yil, ders_kodu, ders_adi,
    toplam_katilimci, tercih_sayisi, aciklama
  - "Meta" sayfasi: aciklama / format notu

Bu format `app/services/survey_import_service.parse_survey_excel` tarafindan
dogrudan okunur (ders_kodu + tercih_sayisi/oy_sayisi + satir bazli fakulte/yil).

Kullanim:
    python -m scripts.generate_anket_veri_seti
    python -m scripts.generate_anket_veri_seti --yil 2022 --katilimci 100
    python -m scripts.generate_anket_veri_seti --db data/adil_secmeli.db
"""
from __future__ import annotations

import argparse
import random
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

KOK = Path(__file__).resolve().parent.parent

SECMELI_ANAHTAR = ("secmeli", "elective")


def _norm(s: str | None) -> str:
    t = str(s or "").strip().lower()
    for a, b in (("ç", "c"), ("ğ", "g"), ("ı", "i"), ("ö", "o"), ("ş", "s"), ("ü", "u")):
        t = t.replace(a, b)
    return t


def _faculty_courses(conn: sqlite3.Connection, faculty_id: int) -> list[tuple[str, str]]:
    """Fakultenin (tercihen secmeli) derslerini (kod, ad) doner."""
    cur = conn.cursor()
    # Ders tipi kolonu varsa secmelileri sec; yoksa tum dersler.
    cur.execute("PRAGMA table_info(ders)")
    cols = {r[1].lower() for r in cur.fetchall()}
    tip_col = next((c for c in ("derstipi", "ders_tipi", "tip", "tur") if c in cols), None)
    cur.execute(
        "SELECT kod, ad, COALESCE(%s,'') FROM ders WHERE fakulte_id = ? AND kod IS NOT NULL"
        % (tip_col or "''"),
        (int(faculty_id),),
    )
    rows = cur.fetchall()
    secmeli = [(r[0], r[1]) for r in rows if any(k in _norm(r[2]) for k in SECMELI_ANAHTAR)]
    # Secmeli bulunamazsa tum dersleri aday yap.
    return secmeli if secmeli else [(r[0], r[1]) for r in rows]


def uret(db_path: Path, yil: int, katilimci: int) -> Path:
    random.seed(2022)  # tekrar uretilebilir
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        cur = conn.cursor()
        cur.execute("SELECT fakulte_id, ad FROM fakulte ORDER BY fakulte_id")
        faculties = [(int(r[0]), str(r[1] or "")) for r in cur.fetchall()]

        rows: list[dict] = []
        for fid, fad in faculties:
            courses = _faculty_courses(conn, fid)
            if not courses:
                continue
            k = min(len(courses), random.randint(10, 20))
            secilen = random.sample(courses, k)
            for kod, ad in secilen:
                # Rastgele tercih: katilimcinin bir kismi bu dersi secmis.
                tercih = random.randint(3, katilimci)
                rows.append({
                    "fakulte_adi": fad,
                    "yil": yil,
                    "ders_kodu": kod,
                    "ders_adi": ad,
                    "toplam_katilimci": katilimci,
                    "tercih_sayisi": tercih,
                    "aciklama": "",
                })
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # Workbook() daima aktif sheet ile gelir
    ws.title = "AnketSonuclari"
    headers = ["fakulte_adi", "yil", "ders_kodu", "ders_adi",
               "toplam_katilimci", "tercih_sayisi", "aciklama"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1565C0")
        c.alignment = Alignment(horizontal="center")
    for ri, r in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=r[h])
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        col_idx = col[0].column
        if col_idx is None:
            continue
        ws.column_dimensions[get_column_letter(int(col_idx))].width = min(max(10, w + 2), 40)
    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Meta")
    meta_satirlar = [
        ("Aciklama", "Universite secmeli ders anketi — ogrenci tercih sayilari"),
        ("yil", yil),
        ("fakulte_basina_katilimci", katilimci),
        ("not", "Her fakultede 10-20 ders anket kapsaminda; tercih sayilari rastgele."),
        ("format", "ders_kodu + tercih_sayisi sistem tarafindan okunur (survey_import_service)."),
    ]
    for ri, (k, v) in enumerate(meta_satirlar, 1):
        meta.cell(row=ri, column=1, value=k).font = Font(bold=True)
        meta.cell(row=ri, column=2, value=v)
    meta.column_dimensions["A"].width = 26
    meta.column_dimensions["B"].width = 60

    cikti = KOK / "data" / f"{yil}_anket_tercih_veri_seti.xlsx"
    wb.save(str(cikti))
    print(f"Kaydedildi: {cikti}")
    print(f"  Toplam anket satiri: {len(rows)}  |  Fakulte sayisi: {len(faculties)}")
    return cikti


def main() -> int:
    parser = argparse.ArgumentParser(description="Anket/tercih veri seti uretici")
    parser.add_argument("--yil", type=int, default=2022)
    parser.add_argument("--katilimci", type=int, default=100)
    parser.add_argument("--db", default=str(KOK / "data" / "adil_secmeli.db"))
    args = parser.parse_args()

    db_path = Path(args.db).resolve()
    if not db_path.exists():
        print(f"HATA: DB bulunamadi: {db_path}", file=sys.stderr)
        return 2
    uret(db_path, args.yil, args.katilimci)
    return 0


if __name__ == "__main__":
    sys.exit(main())
