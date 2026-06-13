import os
import sqlite3
import tempfile

import pandas as pd
from openpyxl import load_workbook

from app.services.survey_import_service import (
    SURVEY_TEMPLATE_SHEET_NAME,
    compute_total_participants,
    import_survey_excel,
    write_survey_template_excel,
)


def _build_db() -> str:
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    conn = sqlite3.connect(path)
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE fakulte (fakulte_id INTEGER PRIMARY KEY, ad TEXT);
        CREATE TABLE bolum (bolum_id INTEGER PRIMARY KEY, fakulte_id INTEGER, ad TEXT);
        CREATE TABLE ders (
            ders_id INTEGER PRIMARY KEY,
            kod TEXT,
            ad TEXT,
            bolum_id INTEGER,
            fakulte_id INTEGER,
            DersTipi TEXT
        );
        CREATE TABLE mufredat (
            mufredat_id INTEGER PRIMARY KEY AUTOINCREMENT,
            fakulte_id INTEGER,
            akademik_yil INTEGER,
            bolum_id INTEGER,
            donem TEXT,
            durum TEXT,
            versiyon INTEGER
        );
        CREATE TABLE mufredat_ders (
            mders_id INTEGER PRIMARY KEY AUTOINCREMENT,
            mufredat_id INTEGER,
            ders_id INTEGER
        );
        CREATE TABLE havuz (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ders_id TEXT,
            yil INTEGER,
            fakulte_id INTEGER,
            bolum_id INTEGER,
            donem TEXT,
            statu INTEGER,
            sayac INTEGER,
            skor REAL,
            ders_adi TEXT
        );
        CREATE TABLE ders_kriterleri (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ders_id INTEGER,
            yil INTEGER,
            donem TEXT,
            toplam_ogrenci INTEGER DEFAULT 0,
            gecen_ogrenci INTEGER DEFAULT 0,
            basari_ortalamasi REAL DEFAULT 0.0,
            kontenjan INTEGER DEFAULT 0,
            kayitli_ogrenci INTEGER DEFAULT 0,
            aktif_mi INTEGER DEFAULT 1,
            anket_katilimci INTEGER DEFAULT 0,
            anket_dersi_secen INTEGER DEFAULT 0
        );
        """
    )
    cur.execute("INSERT INTO fakulte VALUES (1, 'Guzel Sanatlar Fakultesi')")
    cur.execute("INSERT INTO bolum VALUES (10, 1, 'Gastronomi')")
    cur.executemany(
        """
        INSERT INTO ders (ders_id, kod, ad, bolum_id, fakulte_id, DersTipi)
        VALUES (?, ?, ?, 10, 1, 'Secmeli')
        """,
        [
            (101, "GST101", "Antik Mutfaklar"),
            (102, "GST102", "Modern Sunum Teknikleri"),
            (103, "GST103", "Gorsel Tabak Tasarimi"),
        ],
    )
    cur.execute(
        """
        INSERT INTO mufredat (fakulte_id, akademik_yil, bolum_id, donem, durum, versiyon)
        VALUES (1, 2022, 10, 'Guz', 'Resmi', 1)
        """
    )
    mid = int(cur.lastrowid)
    cur.executemany(
        "INSERT INTO mufredat_ders (mufredat_id, ders_id) VALUES (?, ?)",
        [(mid, 101), (mid, 102), (mid, 103)],
    )
    cur.executemany(
        """
        INSERT INTO havuz (ders_id, yil, fakulte_id, bolum_id, donem, statu, sayac, skor, ders_adi)
        VALUES (?, 2022, 1, 10, 'Guz', 1, 0, 50.0, ?)
        """,
        [
            ("101", "Antik Mutfaklar"),
            ("102", "Modern Sunum Teknikleri"),
            ("103", "Gorsel Tabak Tasarimi"),
        ],
    )
    cur.executemany(
        """
        INSERT INTO ders_kriterleri
            (ders_id, yil, donem, toplam_ogrenci, gecen_ogrenci, basari_ortalamasi,
             kontenjan, kayitli_ogrenci, aktif_mi, anket_katilimci, anket_dersi_secen)
        VALUES (?, 2022, 'Güz', 80, 60, 72.0, 40, 80, 1, ?, ?)
        """,
        [
            (101, 50, 20),
            (102, 50, 10),
            (103, 50, 5),
        ],
    )
    conn.commit()
    conn.close()
    return path


def _build_legacy_db_without_kod() -> str:
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    conn = sqlite3.connect(path)
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE fakulte (fakulte_id INTEGER PRIMARY KEY, ad TEXT);
        CREATE TABLE bolum (bolum_id INTEGER PRIMARY KEY, fakulte_id INTEGER, ad TEXT);
        CREATE TABLE ders (
            ders_id INTEGER PRIMARY KEY,
            ad TEXT,
            bolum_id INTEGER,
            fakulte_id INTEGER,
            DersTipi TEXT
        );
        CREATE TABLE havuz (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ders_id TEXT,
            yil INTEGER,
            fakulte_id INTEGER,
            bolum_id INTEGER,
            donem TEXT,
            statu INTEGER,
            sayac INTEGER,
            skor REAL,
            ders_adi TEXT
        );
        """
    )
    cur.execute("INSERT INTO fakulte VALUES (1, 'Guzel Sanatlar Fakultesi')")
    cur.execute("INSERT INTO bolum VALUES (10, 1, 'Gastronomi')")
    cur.executemany(
        """
        INSERT INTO ders (ders_id, ad, bolum_id, fakulte_id, DersTipi)
        VALUES (?, ?, 10, 1, 'Secmeli')
        """,
        [
            (101, "Antik Mutfaklar"),
            (102, "Modern Sunum Teknikleri"),
        ],
    )
    cur.executemany(
        """
        INSERT INTO havuz (ders_id, yil, fakulte_id, bolum_id, donem, statu, sayac, skor, ders_adi)
        VALUES (?, 2022, 1, 10, 'Guz', ?, 0, 50.0, ?)
        """,
        [
            ("101", 1, "Antik Mutfaklar"),
            ("102", 0, "Modern Sunum Teknikleri"),
        ],
    )
    conn.commit()
    conn.close()
    return path


def _write_survey_excel(meta: dict, rows: list[dict]) -> str:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([meta]).to_excel(writer, sheet_name="Meta", index=False)
        pd.DataFrame(rows).to_excel(writer, sheet_name="AnketSonuclari", index=False)
    return path


def _write_flat_survey_excel(rows: list[dict], sheet_name: str = SURVEY_TEMPLATE_SHEET_NAME) -> str:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)
    return path


def test_import_survey_replaces_previous_data():
    db_path = _build_db()
    excel_path_1 = _write_survey_excel(
        {"fakulte_adi": "Guzel Sanatlar Fakultesi", "yil": 2022, "toplam_katilimci": 100},
        [
            {"ders_kodu": "GST101", "ders_adi": "Antik Mutfaklar", "tercih_sayisi": 40},
            {"ders_kodu": "GST102", "ders_adi": "Modern Sunum Teknikleri", "tercih_sayisi": 60},
        ],
    )
    excel_path_2 = _write_survey_excel(
        {"fakulte_adi": "Guzel Sanatlar Fakultesi", "yil": 2022, "toplam_katilimci": 100},
        [
            {"ders_kodu": "GST101", "ders_adi": "Antik Mutfaklar", "tercih_sayisi": 30},
            {"ders_adi": "Gorsel Tabak Tasarimi", "tercih_sayisi": 70},
        ],
    )
    try:
        first = import_survey_excel(db_path=db_path, excel_path=excel_path_1, faculty_id=1, year=2022)
        second = import_survey_excel(db_path=db_path, excel_path=excel_path_2, faculty_id=1, year=2022)

        assert first["ok"] is True
        assert second["ok"] is True
        assert second["replace"]["previous_import_deleted"] == 1

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM survey_import WHERE fakulte_id = 1 AND yil = 2022")
        assert int(cur.fetchone()[0] or 0) == 1
        cur.execute(
            """
            SELECT ders_id, anket_katilimci, anket_dersi_secen
            FROM ders_kriterleri
            WHERE yil = 2022
            ORDER BY ders_id
            """
        )
        rows = cur.fetchall()
        conn.close()

        assert rows == [
            (101, 100, 30),
            (102, 100, 0),
            (103, 100, 70),
        ]
    finally:
        for path in (db_path, excel_path_1, excel_path_2):
            try:
                os.unlink(path)
            except OSError:
                pass


def test_total_participants_computed():
    total = compute_total_participants(
        [
            type("Row", (), {"tercih_sayisi": 15})(),
            type("Row", (), {"tercih_sayisi": 35})(),
            type("Row", (), {"tercih_sayisi": 50})(),
        ]
    )
    assert total == 100


def test_write_survey_template_prefills_active_pool_courses():
    db_path = _build_db()
    fd, template_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.executemany(
        """
        INSERT INTO ders (ders_id, kod, ad, bolum_id, fakulte_id, DersTipi)
        VALUES (?, ?, ?, 10, 1, ?)
        """,
        [
            (104, "GST104", "Deneysel Tipografi", "Secmeli"),
            (105, "GST105", "Kalici Kapanan Ders", "Secmeli"),
            (106, "GST106", "Dinlenen Ders", "Secmeli"),
            (107, "GST107", "Zorunlu Atolye", "Zorunlu"),
        ],
    )
    cur.executemany(
        """
        INSERT INTO havuz (ders_id, yil, fakulte_id, bolum_id, donem, statu, sayac, skor, ders_adi)
        VALUES (?, 2022, 1, 10, 'Guz', ?, 0, 50.0, ?)
        """,
        [
            ("104", 0, "Deneysel Tipografi"),
            ("105", -2, "Kalici Kapanan Ders"),
            ("106", -1, "Dinlenen Ders"),
            ("107", 1, "Zorunlu Atolye"),
        ],
    )
    conn.commit()
    conn.close()

    try:
        write_survey_template_excel(
            target_path=template_path,
            faculty_name="Guzel Sanatlar Fakultesi",
            year=2022,
            db_path=db_path,
            faculty_id=1,
        )

        survey_df = pd.read_excel(template_path, sheet_name=SURVEY_TEMPLATE_SHEET_NAME, dtype=object)
        workbook = load_workbook(template_path, data_only=False)
        worksheet = workbook[SURVEY_TEMPLATE_SHEET_NAME]

        assert survey_df["fakulte_adi"].dropna().unique().tolist() == ["Guzel Sanatlar Fakultesi"]
        assert survey_df["yil"].dropna().astype(int).unique().tolist() == [2022]
        assert set(survey_df["ders_kodu"].dropna().tolist()) == {"GST101", "GST102", "GST103", "GST104"}
        assert "GST105" not in set(survey_df["ders_kodu"].dropna().tolist())
        assert "GST106" not in set(survey_df["ders_kodu"].dropna().tolist())
        assert "GST107" not in set(survey_df["ders_kodu"].dropna().tolist())
        # Kolon duzeni: fakulte_adi|yil|donem|ders_kodu|ders_adi|toplam_katilimci|tercih_sayisi
        # ders_adi -> E sutunu, tercih_sayisi -> G sutunu
        assert worksheet["E6"].value == "TOPLAM"
        assert worksheet["G6"].value == "=SUM(G2:G5)"
    finally:
        for path in (db_path, template_path):
            try:
                os.unlink(path)
            except OSError:
                pass


def test_write_survey_template_handles_legacy_ders_schema_without_kod():
    db_path = _build_legacy_db_without_kod()
    fd, template_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    try:
        write_survey_template_excel(
            target_path=template_path,
            year=2022,
            db_path=db_path,
            faculty_id=1,
        )

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(ders)")
        columns = {row[1] for row in cur.fetchall()}
        conn.close()

        survey_df = pd.read_excel(template_path, sheet_name=SURVEY_TEMPLATE_SHEET_NAME, dtype=object)
        workbook = load_workbook(template_path, data_only=False)
        worksheet = workbook[SURVEY_TEMPLATE_SHEET_NAME]

        assert "kod" in columns
        assert survey_df["ders_adi"].dropna().tolist() == [
            "Antik Mutfaklar",
            "Modern Sunum Teknikleri",
            "TOPLAM",
        ]
        # tercih_sayisi -> G sutunu (yeni kolon duzeni)
        assert worksheet["G4"].value == "=SUM(G2:G3)"
    finally:
        for path in (db_path, template_path):
            try:
                os.unlink(path)
            except OSError:
                pass


def test_import_survey_supports_flat_template_format():
    db_path = _build_db()
    excel_path = _write_flat_survey_excel(
        [
            {
                "fakulte_adi": "Guzel Sanatlar Fakultesi",
                "yil": 2022,
                "ders_kodu": "GST101",
                "ders_adi": "Antik Mutfaklar",
                "oy_miktari": 35,
            },
            {
                "fakulte_adi": "Guzel Sanatlar Fakultesi",
                "yil": 2022,
                "ders_kodu": "GST102",
                "ders_adi": "Modern Sunum Teknikleri",
                "oy_miktari": 65,
            },
        ]
    )
    try:
        result = import_survey_excel(db_path=db_path, excel_path=excel_path, faculty_id=1, year=2022)
        assert result["ok"] is True

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT ders_id, anket_dersi_secen
            FROM ders_kriterleri
            WHERE yil = 2022 AND ders_id IN (101, 102)
            ORDER BY ders_id
            """
        )
        rows = cur.fetchall()
        conn.close()

        assert rows == [(101, 35), (102, 65)]
    finally:
        for path in (db_path, excel_path):
            try:
                os.unlink(path)
            except OSError:
                pass


def test_import_survey_ignores_template_total_row():
    db_path = _build_db()
    excel_path = _write_flat_survey_excel(
        [
            {
                "fakulte_adi": "Guzel Sanatlar Fakultesi",
                "yil": 2022,
                "ders_kodu": "GST101",
                "ders_adi": "Antik Mutfaklar",
                "oy_miktari": 35,
            },
            {
                "fakulte_adi": "Guzel Sanatlar Fakultesi",
                "yil": 2022,
                "ders_kodu": "GST102",
                "ders_adi": "Modern Sunum Teknikleri",
                "oy_miktari": 65,
            },
            {
                "fakulte_adi": None,
                "yil": None,
                "ders_kodu": None,
                "ders_adi": "TOPLAM",
                "oy_miktari": 100,
            },
        ]
    )
    try:
        result = import_survey_excel(db_path=db_path, excel_path=excel_path, faculty_id=1, year=2022)
        assert result["ok"] is True
        assert result["matched_count"] == 2

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT ders_id, anket_dersi_secen
            FROM ders_kriterleri
            WHERE yil = 2022 AND ders_id IN (101, 102)
            ORDER BY ders_id
            """
        )
        rows = cur.fetchall()
        conn.close()

        assert rows == [(101, 35), (102, 65)]
    finally:
        for path in (db_path, excel_path):
            try:
                os.unlink(path)
            except OSError:
                pass


def test_course_matching_by_code_then_name():
    db_path = _build_db()
    excel_path = _write_survey_excel(
        {"fakulte_adi": "Guzel Sanatlar Fakultesi", "yil": 2022, "toplam_katilimci": 100},
        [
            {"ders_kodu": "GST101", "tercih_sayisi": 45},
            {"ders_adi": "gorsel tabak tasarimi", "tercih_sayisi": 55},
        ],
    )
    try:
        result = import_survey_excel(db_path=db_path, excel_path=excel_path, faculty_id=1, year=2022)
        assert result["ok"] is True

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT ders_id, anket_dersi_secen
            FROM ders_kriterleri
            WHERE yil = 2022 AND ders_id IN (101, 103)
            ORDER BY ders_id
            """
        )
        rows = cur.fetchall()
        conn.close()

        assert rows == [(101, 45), (103, 55)]
    finally:
        for path in (db_path, excel_path):
            try:
                os.unlink(path)
            except OSError:
                pass


def test_unmatched_rows_reported():
    db_path = _build_db()
    excel_path = _write_survey_excel(
        {"fakulte_adi": "Guzel Sanatlar Fakultesi", "yil": 2022, "toplam_katilimci": 100},
        [
            {"ders_kodu": "GST101", "tercih_sayisi": 40},
            {"ders_adi": "Sistemde Olmayan Ders", "tercih_sayisi": 60},
        ],
    )
    try:
        result = import_survey_excel(db_path=db_path, excel_path=excel_path, faculty_id=1, year=2022)

        assert result["ok"] is False
        assert result["matched_count"] == 1
        assert result["unmatched_count"] == 1
        assert any("Sistemde eslesen ders bulunamadi" in err for err in result["errors"])

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM survey_import")
        assert int(cur.fetchone()[0] or 0) == 0
        conn.close()
    finally:
        for path in (db_path, excel_path):
            try:
                os.unlink(path)
            except OSError:
                pass


def test_survey_values_written_to_criteria():
    db_path = _build_db()
    excel_path = _write_survey_excel(
        {"fakulte_adi": "Guzel Sanatlar Fakultesi", "yil": 2022, "toplam_katilimci": 100},
        [
            {"ders_kodu": "GST101", "tercih_sayisi": 25},
            {"ders_kodu": "GST102", "tercih_sayisi": 75},
        ],
    )
    try:
        result = import_survey_excel(db_path=db_path, excel_path=excel_path, faculty_id=1, year=2022)
        assert result["ok"] is True

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT anket_katilimci, anket_dersi_secen, anket_veri_kaynagi,
                   anket_manual_locked, anket_import_id
            FROM ders_kriterleri
            WHERE ders_id = 101 AND yil = 2022
            """
        )
        row = cur.fetchone()
        conn.close()

        assert row == (100, 25, "survey_import", 1, result["import_id"])
    finally:
        for path in (db_path, excel_path):
            try:
                os.unlink(path)
            except OSError:
                pass
