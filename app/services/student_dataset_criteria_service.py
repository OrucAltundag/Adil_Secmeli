# -*- coding: utf-8 -*-
"""
Ogrenci not veri setinden OTOMATIK kriter uretici.

Kaynak: data/<yil>_ogrenci_not_veri_seti.xlsx -> 'Ders Analizi' sekmesi
Hedef (manuel kayit ile AYNI uc tablo):
  - ders_kriterleri
  - performans   (ortalama_not, basari_orani)
  - populerlik   (talep_sayisi, kontenjan, doluluk_orani)

Dersler `ders.kod` ile eslesir; eslesmeyen ders icin yeni ders olusturulmaz.
Manuel "Verileri Kaydet" islemi zaten bu uc tabloyu birden yazar; bu fonksiyon
da tutarlilik icin ayni uc tabloyu doldurur (aksi halde Veri Kalitesi'nde
performans/populerlik %0 gorunur).
"""
from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from app.db.schema_compat import ensure_criteria_import_schema
from app.services.popularity_service import calculate_popularity_score

DATA_DIR = Path(__file__).parent.parent.parent / "data"
YIL = 2022
VARSAYILAN_EXCEL = DATA_DIR / f"{YIL}_ogrenci_not_veri_seti.xlsx"


def _dataset_for_year(year: int) -> Path:
    """Yila uygun not veri seti dosyasini doner (data/<yil>_...xlsx)."""
    return DATA_DIR / f"{int(year)}_ogrenci_not_veri_seti.xlsx"


def _read_ders_analizi(excel_path: str | Path) -> list[dict[str, Any]]:
    """Ogrenci not veri setinin 'Ders Analizi' sayfasini okur (ham kayitlar).

    Sadece okuma yapar; hicbir DB erisimi/yazma yoktur. Kolon eksikse hata verir.
    """
    yol = Path(excel_path)
    if not yol.exists():
        raise FileNotFoundError(f"Ogrenci veri seti bulunamadi: {yol}")
    wb = openpyxl.load_workbook(str(yol), read_only=True)
    if "Ders Analizi" not in wb.sheetnames:
        wb.close()
        raise ValueError("Excel dosyasinda 'Ders Analizi' sekmesi yok.")
    ws = wb["Ders Analizi"]
    it = ws.iter_rows(min_row=1, values_only=True)
    hdr = list(next(it))
    j = {str(k): i for i, k in enumerate(hdr)}
    gerekli = {"ders_kodu", "donem", "kayit_sayisi", "gecme_orani_%", "ort_agirlikli", "ort_katilim_yuzde"}
    if not gerekli.issubset(set(j.keys())):
        wb.close()
        raise ValueError(f"Excel 'Ders Analizi' sekmesinde gerekli sutunlar eksik: {gerekli - set(j.keys())}")
    kayitlar: list[dict[str, Any]] = []
    for r in it:
        kayitlar.append({
            "kod": str(r[j["ders_kodu"]]).strip(),
            "donem": str(r[j["donem"]]).strip(),
            "kayit": int(r[j["kayit_sayisi"]] or 0),  # type: ignore[arg-type]
            "gecme": float(r[j["gecme_orani_%"]] or 0),  # type: ignore[arg-type]
            "agir": float(r[j["ort_agirlikli"]] or 0),  # type: ignore[arg-type]
            "katilim": float(r[j["ort_katilim_yuzde"]] or 0),  # type: ignore[arg-type]
        })

    # Dört katılım alanı Ana Veri'de öğrenci bazındadır. Kriter dosyası ders
    # bazında tek satır tuttuğu için ortalama katılım sayısı/yüzdesi, toplam
    # hafta ve devamsız öğrenci sayısı olarak güvenli biçimde özetlenir.
    attendance_by_course: dict[tuple[str, str], dict[str, float]] = {}
    if "Ana Veri" in wb.sheetnames:
        raw_ws = wb["Ana Veri"]
        raw_it = raw_ws.iter_rows(min_row=1, values_only=True)
        raw_header = [str(value or "").strip() for value in next(raw_it)]
        raw_index = {name: idx for idx, name in enumerate(raw_header)}
        attendance_fields = {
            "ders_kodu",
            "donem",
            "katilim_sayisi",
            "toplam_hafta",
            "katilim_yuzdesi",
            "devamsiz_mi",
        }
        if attendance_fields.issubset(raw_index):
            for raw_row in raw_it:
                code = str(raw_row[raw_index["ders_kodu"]] or "").strip()
                term = str(raw_row[raw_index["donem"]] or "").strip().lower()[:1]
                if not code:
                    continue
                key = (code, term)
                aggregate = attendance_by_course.setdefault(
                    key,
                    {
                        "count": 0.0,
                        "attendance_sum": 0.0,
                        "percentage_sum": 0.0,
                        "total_weeks": 0.0,
                        "absent_count": 0.0,
                    },
                )
                try:
                    attendance_count = float(raw_row[raw_index["katilim_sayisi"]] or 0)
                except (TypeError, ValueError):
                    attendance_count = 0.0
                try:
                    attendance_percentage = float(raw_row[raw_index["katilim_yuzdesi"]] or 0)
                except (TypeError, ValueError):
                    attendance_percentage = 0.0
                try:
                    total_weeks = float(raw_row[raw_index["toplam_hafta"]] or 0)
                except (TypeError, ValueError):
                    total_weeks = 0.0
                absent_text = str(raw_row[raw_index["devamsiz_mi"]] or "").strip().lower()
                aggregate["count"] += 1.0
                aggregate["attendance_sum"] += attendance_count
                aggregate["percentage_sum"] += attendance_percentage
                aggregate["total_weeks"] = max(aggregate["total_weeks"], total_weeks)
                if absent_text.startswith(("e", "y", "t")):
                    aggregate["absent_count"] += 1.0

    for record in kayitlar:
        key = (str(record["kod"]), str(record["donem"] or "").strip().lower()[:1])
        aggregate = attendance_by_course.get(key)
        if aggregate and aggregate["count"] > 0:
            count = aggregate["count"]
            record["katilim_sayisi"] = round(aggregate["attendance_sum"] / count, 2)
            record["toplam_hafta"] = int(aggregate["total_weeks"])
            record["katilim_yuzdesi"] = round(aggregate["percentage_sum"] / count, 2)
            record["devamsiz_ogrenci_sayisi"] = int(aggregate["absent_count"])
        else:
            # Eski dosyalarda Ana Veri yoksa Ders Analizi yüzdesini koru.
            record["katilim_sayisi"] = None
            record["toplam_hafta"] = None
            record["katilim_yuzdesi"] = round(float(record["katilim"]), 2)
            record["devamsiz_ogrenci_sayisi"] = None
    wb.close()
    return kayitlar


def build_student_criteria_dataset(
    excel_path: str | Path | None = None,
    year: int = YIL,
) -> list[dict[str, Any]]:
    """§3: Ogrenci not veri setinden KRITER IMPORT SABLONU formatinda satirlar uretir.

    Cikti dogrudan veritabanina yazilmaz; indirilebilir Excel'e donusturulup
    normal kriter import akisindan (onayli) iceri alinabilir. Kolonlar criteria
    import sablonuyla uyumludur.
    """
    yol = Path(excel_path) if excel_path else _dataset_for_year(year)
    kayitlar = _read_ders_analizi(yol)
    KONTENJAN = 60  # veri setinde kontenjan yok; manuel kayitla ayni varsayilan
    rows: list[dict[str, Any]] = []
    for s in kayitlar:
        if not s["kod"]:
            continue
        toplam = int(s["kayit"])
        gecen = round(toplam * s["gecme"] / 100.0)
        rows.append({
            "ders_kodu": s["kod"],
            "donem": s["donem"],
            "toplam_ogrenci": toplam,
            "gecen_ogrenci": gecen,
            "basari_ortalamasi": round(s["agir"], 2),
            "kontenjan": KONTENJAN,
            "kayitli_ogrenci": toplam,
            "katilim_sayisi": s.get("katilim_sayisi"),
            "toplam_hafta": s.get("toplam_hafta"),
            "katilim_yuzdesi": s.get("katilim_yuzdesi"),
            "devamsiz_ogrenci_sayisi": s.get("devamsiz_ogrenci_sayisi"),
        })
    return rows


def auto_generate_criteria_from_student_dataset(
    conn: sqlite3.Connection,
    *,
    excel_path: str | Path | None = None,
    year: int = YIL,
    replace: bool = True,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Ogrenci veri setinin 'Ders Analizi' sekmesinden ders_kriterleri uretir.

    Args:
        conn: sqlite baglantisi
        excel_path: Excel yolu (yoksa varsayilan)
        year: hedef yil (varsayilan 2022)
        replace: True ise once o yilin satirlarini siler
        dry_run: True ise HICBIR yazma yapmaz (DELETE/INSERT/commit yok); yalniz
            eslesme/onizleme hesaplar. UI onay diyalogunda kullanilir.

    Returns:
        {'eklenen': int, 'eslesmeyen': list[str], 'toplam': int,
         'excel_path': str, 'replace': bool, 'preview_rows': list[dict]}
    """
    yol = Path(excel_path) if excel_path else _dataset_for_year(year)
    if not yol.exists():
        raise FileNotFoundError(f"Ogrenci veri seti bulunamadi: {yol}")

    kayitlar = _read_ders_analizi(yol)

    ensure_criteria_import_schema(conn, commit=False)
    cur = conn.cursor()
    criteria_cols = {
        str(row[1])
        for row in cur.execute("PRAGMA table_info(ders_kriterleri)").fetchall()
    }
    if replace and not dry_run:
        # Manuel kayit ile tutarli: yilin uc tablosunu da temizle.
        cur.execute("DELETE FROM ders_kriterleri WHERE yil = ?", (int(year),))
        cur.execute("DELETE FROM performans WHERE akademik_yil = ?", (int(year),))
        cur.execute("DELETE FROM populerlik WHERE akademik_yil = ?", (int(year),))

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    KONTENJAN = 60  # veri setinde kontenjan yok; manuel kayitla ayni varsayilan
    eklenen, perf_yazilan, pop_yazilan, eslesmeyen = 0, 0, 0, []
    yazilan_ders_ids: list[int] = []
    preview_rows: list[dict[str, Any]] = []
    for s in kayitlar:
        if not s["kod"]:
            continue
        cur.execute("SELECT ders_id FROM ders WHERE kod = ?", (s["kod"],))
        row = cur.fetchone()
        if not row:
            eslesmeyen.append(s["kod"])
            continue
        did = int(row[0])
        yazilan_ders_ids.append(did)
        donem = s["donem"]
        gecen = round(s["kayit"] * s["gecme"] / 100.0)
        anket_kat = round(s["kayit"] * s["katilim"] / 100.0)
        # Turetilen olcumler: başarı ayrı, popülerlik kapasite + katılım
        # bileşenlerinden ortak servis ile hesaplanır.
        basari_orani = (gecen / s["kayit"]) if s["kayit"] > 0 else 0.0
        popularity = calculate_popularity_score(
            capacity=KONTENJAN,
            enrolled=s["kayit"],
            attendance_count=s.get("katilim_sayisi"),
            total_weeks=s.get("toplam_hafta"),
            attendance_percentage=s.get("katilim_yuzdesi"),
            absent_student_count=s.get("devamsiz_ogrenci_sayisi"),
        )
        doluluk_orani = float(popularity["occupancy_ratio"] or 0.0)
        populerlik_puani = float(popularity["popularity_score"] or 0.0)

        # Onizleme satiri (UI onay diyalogu icin) — yazma olsun olmasin doldurulur.
        preview_rows.append({
            "kod": s["kod"],
            "donem": donem,
            "kayit": s["kayit"],
            "basari_orani": round(basari_orani, 3),
            "ortalama_not": round(s["agir"], 2),
            "doluluk_orani": round(doluluk_orani, 3),
            "katilim_yuzdesi": s.get("katilim_yuzdesi"),
            "populerlik_orani": round(populerlik_puani, 3),
        })
        eklenen += 1

        if dry_run:
            # Yazma yok: yalniz eslesme ve onizleme sayilari hesaplanir.
            perf_yazilan += 1
            pop_yazilan += 1
            continue

        kriter_values: list[tuple[str, Any]] = [
            ("ders_id", did),
            ("yil", int(year)),
            ("donem", donem),
            ("toplam_ogrenci", s["kayit"]),
            ("gecen_ogrenci", gecen),
            ("basari_ortalamasi", round(s["agir"], 2)),
            ("kontenjan", KONTENJAN),
            ("kayitli_ogrenci", s["kayit"]),
            ("anket_katilimci", anket_kat),
            ("katilim_sayisi", s.get("katilim_sayisi")),
            ("toplam_hafta", s.get("toplam_hafta")),
            ("katilim_yuzdesi", s.get("katilim_yuzdesi")),
            ("devamsiz_ogrenci_sayisi", s.get("devamsiz_ogrenci_sayisi")),
            ("anket_dersi_secen", s["kayit"]),
            ("anket_veri_kaynagi", "ogrenci_veri_seti"),
            ("criteria_veri_kaynagi", "ogrenci_veri_seti"),
            ("criteria_updated_at", now),
            ("is_active", 1),
        ]
        insert_cols = [col for col, _ in kriter_values if col in criteria_cols]
        insert_vals = [value for col, value in kriter_values if col in criteria_cols]
        placeholders = ",".join("?" for _ in insert_cols)
        cur.execute(
            f"INSERT INTO ders_kriterleri ({','.join(insert_cols)}) VALUES ({placeholders})",
            tuple(insert_vals),
        )

        # performans — TOPSIS 'basari' kriterinin okudugu tablo
        cur.execute(
            "DELETE FROM performans WHERE ders_id=? AND akademik_yil=? AND donem=?",
            (did, int(year), donem),
        )
        cur.execute(
            "INSERT INTO performans (ders_id, akademik_yil, donem, ortalama_not, basari_orani) "
            "VALUES (?,?,?,?,?)",
            (did, int(year), donem, round(s["agir"], 2), basari_orani),
        )
        perf_yazilan += 1

        # populerlik — TOPSIS 'populerlik' kriteri + acilabilirlik talep skoru
        cur.execute(
            "DELETE FROM populerlik WHERE ders_id=? AND akademik_yil=? AND donem=?",
            (did, int(year), donem),
        )
        cur.execute(
            "INSERT INTO populerlik "
            "(ders_id, akademik_yil, donem, talep_sayisi, kontenjan, doluluk_orani, ilgi_orani, ham_puan) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (
                did,
                int(year),
                donem,
                s["kayit"],
                KONTENJAN,
                doluluk_orani,
                popularity.get("attendance_component"),
                populerlik_puani,
            ),
        )
        pop_yazilan += 1

    # Kriter yazilan dersler artik DOLU; bu derslere ait BAYAT "zorunlu alan
    # bos" kritik dogrulama uyarilarini temizle (olgunluk skorunu haksizca
    # dusurmesin). Tablo yoksa sessizce gec. (dry_run'da yazma yok.)
    if yazilan_ders_ids and not dry_run:
        try:
            ph = ",".join("?" for _ in yazilan_ders_ids)
            cur.execute(
                f"DELETE FROM criteria_validation_issues "
                f"WHERE year = ? AND course_id IN ({ph})",
                (int(year), *[int(i) for i in yazilan_ders_ids]),
            )
        except sqlite3.OperationalError:
            pass
    if not dry_run:
        conn.commit()
    return {
        "eklenen": eklenen,
        "performans_yazilan": perf_yazilan,
        "populerlik_yazilan": pop_yazilan,
        "eslesmeyen": eslesmeyen,
        "toplam": len(kayitlar),
        "excel_path": str(yol),
        "replace": replace,
        "dry_run": dry_run,
        "preview_rows": preview_rows,
    }
