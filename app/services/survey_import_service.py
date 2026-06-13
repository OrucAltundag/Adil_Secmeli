from __future__ import annotations

import json
import os
import sqlite3
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl.styles import Font

from app.core.config import resolve_sqlite_db_path
from app.db.schema_compat import ensure_reporting_schema, ensure_survey_import_schema
from app.db.sqlite_connection import connect_sqlite, is_database_locked_error
from app.services.course_matcher import (
    load_faculty_course_candidates,
    match_course_row,
    normalize_course_key,
    normalize_course_text,
)
from app.services.course_type import build_elective_predicate
from app.services.import_audit_service import (
    calculate_row_hash,
    create_import_batch,
    extract_excel_metadata,
    link_source_import,
    mark_batch_failed_by_path,
    record_import_issue,
    update_import_status,
)
from app.services.import_diff_service import recalculate_import_diff
from app.services.import_impact_service import recalculate_import_impact
from app.services.import_lineage_service import record_value_source
from app.services.import_quality_service import evaluate_import_quality

SURVEY_TEMPLATE_VERSION = "survey-import-v1"
SURVEY_TEMPLATE_SHEET_NAME = "Ders Veri Giriş Şablonu Oluştur"


@dataclass
class SurveyRow:
    row_no: int
    ders_kodu: str | None
    ders_adi: str | None
    tercih_sayisi: int
    aciklama: str | None = None
    fakulte_adi: str | None = None
    yil: int | None = None
    toplam_katilimci: int | None = None
    donem: str | None = None


@dataclass
class SurveyImportRowResult:
    row_no: int
    ders_kodu: str | None
    ders_adi: str | None
    tercih_sayisi: int
    aciklama: str | None = None
    matched_ders_id: int | None = None
    match_method: str | None = None
    row_status: str = "matched"
    error_message: str | None = None
    raw_faculte: str | None = None
    raw_yil: int | None = None
    donem: str | None = None

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


def _now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _normalize_text(value: str | None) -> str:
    return normalize_course_text(value)


def _find_col(columns: list[str], *candidates: str) -> str | None:
    norm_map = {_normalize_text(col): col for col in columns}
    for cand in candidates:
        key = _normalize_text(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def _parse_year(value: Any) -> int | None:
    try:
        if value is None or pd.isna(value):
            return None
        year = int(float(value))
    except (TypeError, ValueError):
        text = str(value or "").strip()
        digits = "".join(ch for ch in text if ch.isdigit())
        if len(digits) >= 4:
            try:
                year = int(digits[:4])
            except ValueError:
                return None
        else:
            return None
    return year if 1900 <= year <= 2100 else None


def _safe_int(value: Any, default: int | None = None) -> int | None:
    try:
        if value is None or pd.isna(value):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _clean_text(value: Any) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    return text or None


def _normalize_donem(value: Any) -> str | None:
    """Serbest metin donem girdisini 'Güz'/'Bahar' standardina cevirir."""
    text = _clean_text(value)
    if not text:
        return None
    key = _normalize_text(text)
    if key in {"guz", "güz", "fall", "autumn"}:
        return "Güz"
    if key in {"bahar", "spring"}:
        return "Bahar"
    return text


def _resolve_faculty_id_by_name(cur: sqlite3.Cursor, faculty_name: str | None) -> int | None:
    """Belge satirindaki fakulte adini fakulte tablosundaki id ile eslestirir."""
    target = _normalize_text(faculty_name)
    if not target:
        return None
    cur.execute("SELECT fakulte_id, ad FROM fakulte")
    for row in cur.fetchall():
        if row and row[0] is not None and _normalize_text(str(row[1] or "")) == target:
            return int(row[0])
    return None


def _filter_rows_for_faculty(rows: list[SurveyRow], faculty_name: str | None) -> list[SurveyRow]:
    """Belgedeki cok-fakulteli satirlardan yalnizca secili fakulteye ait olanlari tutar.

    Fakulte adi bos olan satirlar (eski tek-fakulteli sablonlar) secili fakulteye ait
    kabul edilir; boylece geriye donuk uyumluluk korunur.
    """
    target = _normalize_text(faculty_name)
    if not target:
        return list(rows)
    kept: list[SurveyRow] = []
    for row in rows:
        row_fac = _normalize_text(row.fakulte_adi)
        if not row_fac or row_fac == target:
            kept.append(row)
    return kept


def _excel_column_letter(index_1_based: int) -> str:
    if index_1_based < 1:
        raise ValueError("Excel kolon indeksi 1 veya daha buyuk olmali.")
    out = ""
    current = int(index_1_based)
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        out = chr(65 + remainder) + out
    return out


def _is_summary_row(ders_kodu: str | None, ders_adi: str | None) -> bool:
    if _clean_text(ders_kodu):
        return False
    normalized_name = normalize_course_key(ders_adi)
    return normalized_name in {"toplam", "geneltoplam"}


def _read_meta_sheet(xls: pd.ExcelFile) -> dict[str, Any]:
    meta_sheet = next((name for name in xls.sheet_names if _normalize_text(str(name)) == "meta"), None)
    if not meta_sheet:
        return {}
    df = pd.read_excel(xls, sheet_name=meta_sheet)
    if df.empty:
        return {}
    df.columns = [str(col).strip() for col in df.columns]
    row = df.iloc[0].to_dict()
    return {
        "fakulte_adi": _clean_text(row.get(_find_col(list(df.columns), "fakulte_adi", "fakulte", "faculty"))),
        "yil": _parse_year(row.get(_find_col(list(df.columns), "yil", "akademik_yil", "year"))),
        "toplam_katilimci": _safe_int(row.get(_find_col(list(df.columns), "toplam_katilimci", "ankete_katilan_toplam_ogrenci", "total_participants"))),
        "aciklama": _clean_text(row.get(_find_col(list(df.columns), "aciklama", "not", "notes"))),
    }


def parse_survey_excel(excel_path: str) -> dict[str, Any]:
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Anket dosyasi bulunamadi: {excel_path}")

    xls = pd.ExcelFile(excel_path)
    meta = _read_meta_sheet(xls)
    data_sheet = next(
        (
            name
            for name in xls.sheet_names
            if _normalize_text(str(name)) in {"anketsonuclari", "anket_sonuclari", "anket", "survey"}
        ),
        None,
    )
    if data_sheet is None:
        non_meta_sheets = [name for name in xls.sheet_names if _normalize_text(str(name)) != "meta"]
        if not non_meta_sheets:
            raise ValueError("Anket veri sayfasi bulunamadi.")
        data_sheet = non_meta_sheets[0]

    df = pd.read_excel(xls, sheet_name=data_sheet)
    df.columns = [str(col).strip() for col in df.columns]

    columns = list(df.columns)
    col_code = _find_col(columns, "ders_kodu", "ders kodu", "kod", "course_code")
    col_name = _find_col(columns, "ders_adi", "ders adi", "ders adı", "course_name")
    col_pref = _find_col(
        columns,
        "tercih_sayisi",
        "tercih sayisi",
        "tercih sayısı",
        "oy_sayisi",
        "oy sayisi",
        "oy_miktari",
        "oy miktari",
        "preference_count",
    )
    col_note = _find_col(columns, "aciklama", "açıklama", "not")
    col_faculty = _find_col(columns, "fakulte_adi", "fakulte", "fakülte")
    col_year = _find_col(columns, "yil", "yıl", "akademik_yil", "akademik yıl", "year")
    col_donem = _find_col(columns, "donem", "dönem", "semester", "term")
    col_total = _find_col(columns, "toplam_katilimci", "toplam katilimci", "toplam katılımcı", "ankete_katilan_toplam_ogrenci")

    if not col_pref:
        raise ValueError("Gerekli kolon bulunamadi: tercih_sayisi / oy_miktari")
    if not (col_code or col_name):
        raise ValueError("Ders tanimlayici gerekli: ders_kodu veya ders_adi")

    rows: list[SurveyRow] = []
    warnings: list[str] = []
    for idx, row in df.iterrows():
        ders_kodu = _clean_text(row.get(col_code)) if col_code else None
        ders_adi = _clean_text(row.get(col_name)) if col_name else None
        tercih_sayisi = _safe_int(row.get(col_pref))
        aciklama = _clean_text(row.get(col_note)) if col_note else None
        raw_faculte = _clean_text(row.get(col_faculty)) if col_faculty else None
        raw_yil = _parse_year(row.get(col_year)) if col_year else None
        raw_donem = _normalize_donem(row.get(col_donem)) if col_donem else None
        total_katilimci = _safe_int(row.get(col_total)) if col_total else None

        row_has_context = any([ders_kodu, ders_adi, aciklama, raw_faculte, raw_yil, total_katilimci])
        if not row_has_context and tercih_sayisi is None:
            continue
        if not row_has_context and tercih_sayisi is not None:
            # Sablondaki toplam/formul satiri gibi veri disi satirlari atla.
            continue
        if _is_summary_row(ders_kodu=ders_kodu, ders_adi=ders_adi):
            continue

        if tercih_sayisi is None:
            warnings.append(f"Satir {int(str(idx)) + 2}: tercih_sayisi/oy_miktari bos veya gecersiz.")
            tercih_sayisi = -1

        rows.append(
            SurveyRow(
                row_no=int(str(idx)) + 2,
                ders_kodu=ders_kodu,
                ders_adi=ders_adi,
                tercih_sayisi=int(tercih_sayisi),
                aciklama=aciklama,
                fakulte_adi=raw_faculte,
                yil=raw_yil,
                toplam_katilimci=total_katilimci,
                donem=raw_donem,
            )
        )

    return {
        "meta": meta,
        "rows": rows,
        "warnings": warnings,
        "sheet_name": data_sheet,
        "template_version": SURVEY_TEMPLATE_VERSION,
    }


def validate_survey_rows(
    rows: list[SurveyRow],
    faculty_name: str | None = None,
    year: int | None = None,
    declared_total_participants: int | None = None,
) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []

    if not rows:
        errors.append("Belge icinde aktarilabilir anket satiri yok.")
        return {"ok": False, "errors": errors, "warnings": warnings, "declared_total_participants": declared_total_participants}

    repeated_totals = {
        int(row.toplam_katilimci)
        for row in rows
        if row.toplam_katilimci is not None
    }
    if declared_total_participants is None and len(repeated_totals) == 1:
        declared_total_participants = next(iter(repeated_totals))
    elif len(repeated_totals) > 1:
        errors.append("Belge satirlarindaki toplam_katilimci degerleri birbiriyle tutarsiz.")

    seen_document_keys: dict[str, int] = {}
    for row in rows:
        if not row.ders_kodu and not row.ders_adi:
            errors.append(f"Satir {row.row_no}: ders_kodu veya ders_adi zorunlu.")
        if row.tercih_sayisi < 0:
            errors.append(f"Satir {row.row_no}: tercih_sayisi negatif veya gecersiz.")
        # Not: cok-fakulteli belgelerde satirlar import oncesi secili fakulteye gore
        # filtrelendiginden burada fakulte uyusmazligi ayrica hata uretmez.
        if year is not None and row.yil is not None and int(row.yil) != int(year):
            errors.append(f"Satir {row.row_no}: belge yili '{row.yil}' secili yil '{year}' ile uyusmuyor.")

        course_key = normalize_course_text(row.ders_kodu) or f"ad:{normalize_course_key(row.ders_adi)}"
        # Ayni ders kodu farkli donemlerde (Güz/Bahar) gecebilir; donemi anahtara ekleyerek
        # gercek tekrarlari yanlis pozitiften ayiriyoruz.
        donem_key = _normalize_text(row.donem) if row.donem else ""
        dedupe_key = f"{donem_key}|{course_key}" if course_key else ""
        if dedupe_key:
            if dedupe_key in seen_document_keys:
                errors.append(
                    f"Belgede ayni ders birden fazla kez geciyor: satir {seen_document_keys[dedupe_key]} ve {row.row_no}."
                )
            else:
                seen_document_keys[dedupe_key] = row.row_no

    return {
        "ok": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "declared_total_participants": declared_total_participants,
    }


def compute_total_participants(rows: list[SurveyImportRowResult] | list[SurveyRow]) -> int:
    return int(sum(max(0, int(getattr(row, "tercih_sayisi", 0))) for row in rows))


def match_courses(
    conn: sqlite3.Connection,
    rows: list[SurveyRow],
    faculty_id: int,
    year: int,
) -> dict[str, Any]:
    cur = conn.cursor()
    candidates = load_faculty_course_candidates(cur=cur, faculty_id=int(faculty_id), year=int(year))

    matched_rows: list[SurveyImportRowResult] = []
    unmatched_rows: list[SurveyImportRowResult] = []
    seen_course_ids: dict[int, int] = {}
    errors: list[str] = []

    for row in rows:
        result = match_course_row(candidates=candidates, ders_kodu=row.ders_kodu, ders_adi=row.ders_adi)
        row_result = SurveyImportRowResult(
            row_no=row.row_no,
            ders_kodu=row.ders_kodu,
            ders_adi=row.ders_adi,
            tercih_sayisi=int(row.tercih_sayisi),
            aciklama=row.aciklama,
            raw_faculte=row.fakulte_adi,
            raw_yil=row.yil,
            donem=row.donem,
        )
        if not result.matched or result.ders_id is None:
            row_result.row_status = "unmatched"
            row_result.error_message = result.error or "Sistemde eslesen ders bulunamadi."
            unmatched_rows.append(row_result)
            continue

        if int(result.ders_id) in seen_course_ids:
            first_row = seen_course_ids[int(result.ders_id)]
            row_result.row_status = "duplicate"
            row_result.error_message = (
                f"Ayni ders belge icinde birden fazla kez eslesti: satir {first_row} ve {row.row_no}."
            )
            unmatched_rows.append(row_result)
            continue

        seen_course_ids[int(result.ders_id)] = row.row_no
        row_result.matched_ders_id = int(result.ders_id)
        row_result.match_method = result.match_method
        matched_rows.append(row_result)

    if unmatched_rows:
        errors.extend(
            [
                f"Satir {row.row_no}: {row.error_message or 'Sistemde eslesen ders bulunamadi.'}"
                for row in unmatched_rows
            ]
        )

    return {
        "ok": len(errors) == 0,
        "matched_rows": matched_rows,
        "unmatched_rows": unmatched_rows,
        "matched_count": len(matched_rows),
        "unmatched_count": len(unmatched_rows),
        "errors": errors,
    }


def _resolve_faculty_name(cur: sqlite3.Cursor, faculty_id: int) -> str | None:
    cur.execute("SELECT ad FROM fakulte WHERE fakulte_id = ? LIMIT 1", (int(faculty_id),))
    row = cur.fetchone()
    return str(row[0] or "") if row else None


def _table_exists(cur: sqlite3.Cursor, table_name: str) -> bool:
    cur.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ? LIMIT 1",
        (str(table_name),),
    )
    return cur.fetchone() is not None


def _load_template_courses(
    cur: sqlite3.Cursor,
    faculty_id: int,
    year: int,
) -> list[dict[str, Any]]:
    if not _table_exists(cur, "havuz") or not _table_exists(cur, "ders"):
        raise ValueError("Anket sablonu icin gereken havuz/ders tablolari bulunamadi.")

    try:
        elective_predicate = build_elective_predicate(cur=cur, alias="d")
    except Exception:
        elective_predicate = "1=1"

    if elective_predicate == "0=1":
        elective_predicate = "1=1"

    code_expr = "NULLIF(TRIM(COALESCE(d.kod, '')), '')"
    name_expr = "COALESCE(NULLIF(TRIM(d.ad), ''), NULLIF(TRIM(h.ders_adi), ''), 'Ders ' || h.ders_id)"

    cur.execute(
        f"""
        SELECT
            d.ders_id,
            {code_expr} AS ders_kodu,
            {name_expr} AS ders_adi,
            MAX(CASE WHEN h.statu = 1 THEN 1 ELSE 0 END) AS statu_oncelik,
            MIN(NULLIF(TRIM(COALESCE(h.donem, '')), '')) AS donem
        FROM havuz h
        JOIN ders d ON CAST(h.ders_id AS INTEGER) = d.ders_id
        WHERE h.fakulte_id = ?
          AND h.yil = ?
          AND h.statu IN (0, 1)
          AND {elective_predicate}
        GROUP BY d.ders_id, {code_expr}, {name_expr}
        ORDER BY statu_oncelik DESC, ders_adi, d.ders_id
        """,
        (int(faculty_id), int(year)),
    )
    return [
        {
            "ders_id": int(row[0]),
            "ders_kodu": _clean_text(row[1]),
            "ders_adi": str(row[2] or "").strip(),
            "donem": _normalize_donem(row[4]),
        }
        for row in cur.fetchall()
        if row and row[0] is not None and str(row[2] or "").strip()
    ]


def load_survey_template_context(
    db_path: str,
    faculty_id: int,
    year: int,
) -> dict[str, Any]:
    resolved_db_path = resolve_sqlite_db_path(db_path)
    if not resolved_db_path.exists():
        raise FileNotFoundError("Veritabani bulunamadi.")

    conn = connect_sqlite(str(resolved_db_path))
    try:
        try:
            ensure_reporting_schema(conn)
        except sqlite3.OperationalError as exc:
            if not is_database_locked_error(exc):
                raise
            conn.rollback()
        cur = conn.cursor()
        faculty_name = _resolve_faculty_name(cur, int(faculty_id))
        if not faculty_name:
            raise ValueError("Secili fakulte bulunamadi.")

        courses = _load_template_courses(cur=cur, faculty_id=int(faculty_id), year=int(year))
        return {
            "faculty_name": faculty_name,
            "year": int(year),
            "courses": courses,
        }
    finally:
        conn.close()


def load_all_faculties_template_context(db_path: str, year: int) -> list[dict[str, Any]]:
    """Tum fakulteler icin sablon ders listelerini dondurur ('Tumu' secimi)."""
    resolved_db_path = resolve_sqlite_db_path(db_path)
    if not resolved_db_path.exists():
        raise FileNotFoundError("Veritabani bulunamadi.")

    conn = connect_sqlite(str(resolved_db_path))
    contexts: list[dict[str, Any]] = []
    try:
        try:
            ensure_reporting_schema(conn)
        except sqlite3.OperationalError as exc:
            if not is_database_locked_error(exc):
                raise
            conn.rollback()
        cur = conn.cursor()
        cur.execute("SELECT fakulte_id, ad FROM fakulte ORDER BY ad")
        faculties = [(int(row[0]), str(row[1] or "")) for row in cur.fetchall() if row and row[0] is not None]
        for fid, fac_name in faculties:
            courses = _load_template_courses(cur=cur, faculty_id=int(fid), year=int(year))
            if courses:
                contexts.append({"faculty_name": fac_name, "year": int(year), "courses": courses})
    finally:
        conn.close()
    return contexts


def _get_faculty_year_course_scope(cur: sqlite3.Cursor, faculty_id: int, year: int) -> set[int]:
    scope_ids: set[int] = set()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name IN ('mufredat', 'mufredat_ders', 'bolum', 'havuz')")
    existing_tables = {str(row[0]) for row in cur.fetchall() if row and row[0]}

    if {"mufredat", "mufredat_ders", "bolum"}.issubset(existing_tables):
        cur.execute(
            """
            SELECT DISTINCT md.ders_id
            FROM mufredat m
            JOIN bolum b ON b.bolum_id = m.bolum_id
            JOIN mufredat_ders md ON md.mufredat_id = m.mufredat_id
            WHERE b.fakulte_id = ?
              AND m.akademik_yil = ?
            """,
            (int(faculty_id), int(year)),
        )
        scope_ids.update(int(row[0]) for row in cur.fetchall() if row and row[0] is not None)

    if "havuz" in existing_tables:
        cur.execute(
            """
            SELECT DISTINCT CAST(ders_id AS INTEGER)
            FROM havuz
            WHERE fakulte_id = ? AND yil = ?
            """,
            (int(faculty_id), int(year)),
        )
        scope_ids.update(int(row[0]) for row in cur.fetchall() if row and row[0] is not None)

    if scope_ids:
        return scope_ids

    try:
        elective_predicate = build_elective_predicate(cur=cur, alias="d")
    except Exception:
        elective_predicate = "1=1"

    cur.execute(
        f"""
        SELECT DISTINCT d.ders_id
        FROM ders d
        WHERE d.fakulte_id = ?
          AND {elective_predicate}
        """,
        (int(faculty_id),),
    )
    scope_ids.update(int(row[0]) for row in cur.fetchall() if row and row[0] is not None)
    return scope_ids


def replace_existing_survey_data(
    conn: sqlite3.Connection,
    faculty_id: int,
    year: int,
    scope_course_ids: set[int] | None = None,
) -> dict[str, int]:
    ensure_survey_import_schema(conn, commit=False)
    cur = conn.cursor()
    faculty_id = int(faculty_id)
    year = int(year)
    scope_course_ids = set(scope_course_ids or [])

    cur.execute(
        "SELECT import_id FROM survey_import WHERE fakulte_id = ? AND yil = ? LIMIT 1",
        (faculty_id, year),
    )
    row = cur.fetchone()
    previous_import_id = int(row[0]) if row and row[0] is not None else None

    if previous_import_id is not None:
        cur.execute(
            "SELECT ders_id FROM ders_kriterleri WHERE yil = ? AND anket_import_id = ?",
            (year, previous_import_id),
        )
        scope_course_ids.update(int(item[0]) for item in cur.fetchall() if item and item[0] is not None)
        cur.execute("SELECT matched_ders_id FROM survey_import_rows WHERE import_id = ?", (previous_import_id,))
        scope_course_ids.update(int(item[0]) for item in cur.fetchall() if item and item[0] is not None)

    if not scope_course_ids:
        scope_course_ids = _get_faculty_year_course_scope(cur, faculty_id, year)

    criteria_rows_reset = 0
    for ders_id in sorted(scope_course_ids):
        cur.execute(
            """
            UPDATE ders_kriterleri
            SET anket_katilimci = 0,
                anket_dersi_secen = 0,
                anket_veri_kaynagi = 'manual',
                anket_manual_locked = 0,
                anket_import_id = NULL,
                anket_imported_at = NULL
            WHERE ders_id = ? AND yil = ?
            """,
            (int(ders_id), year),
        )
        criteria_rows_reset += int(cur.rowcount or 0)

    deleted_row_count = 0
    if previous_import_id is not None:
        # Audit trail icin eski satir kayitlarini koruyoruz. survey_import
        # kaydi UNIQUE(fakulte_id, yil) nedeniyle silinir; satirlar
        # import_batch_id ile diff/rollback raporlarinda okunabilir kalir.
        deleted_row_count = 0
        cur.execute("DELETE FROM survey_import WHERE import_id = ?", (previous_import_id,))

    return {
        "previous_import_deleted": 1 if previous_import_id is not None else 0,
        "previous_rows_deleted": deleted_row_count,
        "criteria_rows_reset": criteria_rows_reset,
    }


def apply_survey_to_criteria(
    conn: sqlite3.Connection,
    faculty_id: int,
    year: int,
    rows: list[SurveyImportRowResult],
    source_filename: str | None = None,
    template_version: str = SURVEY_TEMPLATE_VERSION,
    notes: str | None = None,
    import_batch_id: int | None = None,
    total_participants_override: int | None = None,
) -> dict[str, Any]:
    ensure_survey_import_schema(conn, commit=False)
    cur = conn.cursor()
    faculty_id = int(faculty_id)
    year = int(year)
    # Bildirilen katilimci sayisi verildiyse onu kullan (cok-secimli anket); aksi halde
    # tek-secimli kabul edilip tercih toplami katilimci sayilir.
    total_participants = (
        int(total_participants_override)
        if total_participants_override is not None
        else compute_total_participants(rows)
    )
    now = _now_utc()

    matched_course_ids = {int(row.matched_ders_id) for row in rows if row.matched_ders_id is not None}
    scope_course_ids = _get_faculty_year_course_scope(cur, faculty_id, year) | matched_course_ids
    replace_stats = replace_existing_survey_data(
        conn=conn,
        faculty_id=faculty_id,
        year=year,
        scope_course_ids=scope_course_ids,
    )

    cur.execute(
        """
        INSERT INTO survey_import
            (fakulte_id, yil, total_participants, matched_course_count, unmatched_row_count,
             source_filename, template_version, notes, imported_at, status)
        VALUES (?, ?, ?, ?, 0, ?, ?, ?, ?, 'applied')
        """,
        (
            faculty_id,
            year,
            int(total_participants),
            int(len(matched_course_ids)),
            source_filename,
            template_version,
            notes,
            now,
        ),
    )
    import_id = int(cur.lastrowid or 0)
    if import_batch_id is not None:
        cur.execute(
            "UPDATE survey_import SET import_batch_id = ? WHERE import_id = ?",
            (int(import_batch_id), int(import_id)),
        )

    course_pref_map = {int(ders_id): 0 for ders_id in scope_course_ids}
    course_donem_map: dict[int, str] = {}
    row_id_by_course: dict[int, int] = {}
    for row in rows:
        if row.matched_ders_id is None:
            continue
        course_pref_map[int(row.matched_ders_id)] = int(row.tercih_sayisi)
        if getattr(row, "donem", None):
            course_donem_map[int(row.matched_ders_id)] = str(row.donem)
        cur.execute(
            """
            INSERT INTO survey_import_rows
                (import_id, row_no, ders_kodu, ders_adi, tercih_sayisi, aciklama,
                 matched_ders_id, match_method, row_status, error_message, raw_faculte, raw_yil)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                import_id,
                int(row.row_no),
                row.ders_kodu,
                row.ders_adi,
                int(row.tercih_sayisi),
                row.aciklama,
                int(row.matched_ders_id),
                row.match_method,
                row.row_status,
                row.error_message,
                row.raw_faculte,
                row.raw_yil,
            ),
        )
        import_row_id = int(cur.lastrowid or 0)
        row_id_by_course[int(row.matched_ders_id)] = import_row_id
        if import_batch_id is not None:
            normalized_row = asdict(row)
            cur.execute(
                """
                UPDATE survey_import_rows
                SET import_batch_id = ?, normalized_row_json = ?, row_hash = ?
                WHERE row_id = ?
                """,
                (
                    int(import_batch_id),
                    json.dumps(normalized_row, ensure_ascii=False, sort_keys=True, default=str),
                    calculate_row_hash(normalized_row),
                    int(import_row_id),
                ),
            )
            if row.error_message:
                record_import_issue(
                    conn,
                    import_batch_id=int(import_batch_id),
                    source_row_id=int(import_row_id),
                    row_number=int(row.row_no),
                    message=str(row.error_message),
                )
                cur.execute(
                    "UPDATE survey_import_rows SET issue_count = COALESCE(issue_count, 0) + 1 WHERE row_id = ?",
                    (int(import_row_id),),
                )

    created_rows = 0
    updated_rows = 0
    for ders_id in sorted(scope_course_ids):
        tercih_sayisi = int(course_pref_map.get(int(ders_id), 0))
        row_donem = course_donem_map.get(int(ders_id))
        # Belgede donem belirtilmisse o donemdeki kriter satirini tercih et; yoksa en guncel satir.
        if row_donem:
            cur.execute(
                """
                SELECT id
                FROM ders_kriterleri
                WHERE ders_id = ? AND yil = ? AND donem = ?
                ORDER BY id DESC
                LIMIT 1
                """,
                (int(ders_id), year, row_donem),
            )
            existing = cur.fetchone()
        else:
            existing = None
        if existing is None:
            cur.execute(
                """
                SELECT id
                FROM ders_kriterleri
                WHERE ders_id = ? AND yil = ?
                ORDER BY id DESC
                LIMIT 1
                """,
                (int(ders_id), year),
            )
            existing = cur.fetchone()
        if existing:
            cur.execute(
                """
                UPDATE ders_kriterleri
                SET anket_katilimci = ?,
                    anket_dersi_secen = ?,
                    anket_veri_kaynagi = 'survey_import',
                    anket_manual_locked = 1,
                    anket_import_id = ?,
                    anket_imported_at = ?
                WHERE id = ?
                """,
                (total_participants, tercih_sayisi, import_id, now, int(existing[0])),
            )
            updated_rows += int(cur.rowcount or 0)
        else:
            cur.execute(
                """
                INSERT INTO ders_kriterleri
                    (ders_id, yil, donem, toplam_ogrenci, gecen_ogrenci, basari_ortalamasi,
                     kontenjan, kayitli_ogrenci, anket_katilimci, anket_dersi_secen,
                     anket_veri_kaynagi, anket_manual_locked, anket_import_id, anket_imported_at)
                VALUES (?, ?, ?, 0, 0, 0.0, 0, 0, ?, ?, 'survey_import', 1, ?, ?)
                """,
                (int(ders_id), year, row_donem or "Güz", total_participants, tercih_sayisi, import_id, now),
            )
            created_rows += 1
        if import_batch_id is not None:
            source_row_id = row_id_by_course.get(int(ders_id))
            record_value_source(
                conn=conn,
                course_id=int(ders_id),
                year=int(year),
                field_name="anket_katilimci",
                value=total_participants,
                source_type="survey_import",
                faculty_id=int(faculty_id),
                source_import_batch_id=int(import_batch_id),
                source_row_id=source_row_id,
                is_locked=True,
                deactivate_existing=True,
            )
            record_value_source(
                conn=conn,
                course_id=int(ders_id),
                year=int(year),
                field_name="anket_dersi_secen",
                value=tercih_sayisi,
                source_type="survey_import",
                faculty_id=int(faculty_id),
                source_import_batch_id=int(import_batch_id),
                source_row_id=source_row_id,
                is_locked=True,
                deactivate_existing=True,
            )

    return {
        "ok": True,
        "import_id": import_id,
        "total_participants": total_participants,
        "scope_course_count": len(scope_course_ids),
        "created_rows": created_rows,
        "updated_rows": updated_rows,
        "replace": replace_stats,
    }


def _import_survey_all_faculties(
    db_path: str,
    excel_path: str,
    year: int,
    source_filename: str | None = None,
    auto_activate: bool = True,
    uploaded_by: str | None = None,
) -> dict[str, Any]:
    """Belgede gecen tum fakulteleri ayri ayri import eder (Fakulte = 'Tumu' secimi).

    Her fakulte icin import_survey_excel tek-fakulte yoluyla calistirilir ve sonuclar
    birlestirilerek tek bir ozet dondurulur.
    """
    resolved_db_path = resolve_sqlite_db_path(db_path)
    try:
        parsed = parse_survey_excel(excel_path)
    except Exception as exc:
        return {"ok": False, "message": f"Anket dosyasi okunamadi: {exc}", "errors": [str(exc)]}

    rows = list(parsed.get("rows") or [])
    distinct_names = sorted({
        str(row.fakulte_adi).strip()
        for row in rows
        if _clean_text(row.fakulte_adi)
    })
    if not distinct_names:
        return {
            "ok": False,
            "message": "Belgede fakulte bilgisi yok. 'Tumu' secimi icin fakulte_adi kolonu gerekli.",
            "errors": ["Belgede fakulte_adi kolonu bulunamadi veya bos."],
        }

    conn = connect_sqlite(str(resolved_db_path), row_factory=True)
    try:
        cur = conn.cursor()
        name_to_id: dict[str, int] = {}
        unresolved: list[str] = []
        for name in distinct_names:
            fid = _resolve_faculty_id_by_name(cur, name)
            if fid is None:
                unresolved.append(name)
            else:
                name_to_id[name] = fid
    finally:
        conn.close()

    faculty_results: list[dict[str, Any]] = []
    warnings: list[str] = [f"Belgedeki fakulte sistemde bulunamadi: {name}" for name in unresolved]
    errors: list[str] = []
    total_matched = 0
    total_unmatched = 0
    total_participants = 0
    first_batch_id: int | None = None
    success_faculties = 0

    for name in distinct_names:
        fid = name_to_id.get(name)
        if fid is None:
            continue
        result = import_survey_excel(
            db_path=db_path,
            excel_path=excel_path,
            faculty_id=int(fid),
            year=int(year),
            source_filename=source_filename,
            auto_activate=auto_activate,
            uploaded_by=uploaded_by,
        )
        faculty_results.append({"fakulte_adi": name, "faculty_id": int(fid), "result": result})
        total_matched += int(result.get("matched_count") or 0)
        total_unmatched += int(result.get("unmatched_count") or 0)
        total_participants += int(result.get("total_participants") or 0)
        if first_batch_id is None and result.get("import_batch_id") is not None:
            first_batch_id = int(result["import_batch_id"])
        if result.get("ok"):
            success_faculties += 1
        else:
            errors.append(f"{name}: {result.get('message') or 'Import basarisiz.'}")
        for warn in result.get("warnings") or []:
            warnings.append(f"{name}: {warn}")

    overall_ok = success_faculties > 0 and not errors
    summary_message = (
        f"{success_faculties}/{len(name_to_id)} fakulte icin anket verisi yuklendi. "
        f"Toplam eslesen ders: {total_matched}."
    )
    if unresolved:
        summary_message += f" Bulunamayan fakulte sayisi: {len(unresolved)}."

    return {
        "ok": overall_ok,
        "message": summary_message,
        "year": int(year),
        "faculty_results": faculty_results,
        "faculty_count": len(name_to_id),
        "success_faculty_count": success_faculties,
        "unresolved_faculties": unresolved,
        "matched_count": total_matched,
        "unmatched_count": total_unmatched,
        "total_participants": total_participants,
        "import_batch_id": first_batch_id,
        "errors": errors,
        "warnings": warnings,
    }


def import_survey_excel(
    db_path: str,
    excel_path: str,
    faculty_id: int | None,
    year: int,
    source_filename: str | None = None,
    auto_activate: bool = True,
    uploaded_by: str | None = None,
) -> dict[str, Any]:
    resolved_db_path = resolve_sqlite_db_path(db_path)
    if not resolved_db_path.exists():
        return {"ok": False, "message": "Veritabani bulunamadi.", "errors": ["Veritabani bulunamadi."]}
    if not os.path.exists(excel_path):
        return {"ok": False, "message": "Anket dosyasi bulunamadi.", "errors": ["Anket dosyasi bulunamadi."]}

    # Fakulte secilmediyse ("Tumu") belgedeki tum fakulteler tek tek islenir.
    if faculty_id is None:
        return _import_survey_all_faculties(
            db_path=db_path,
            excel_path=excel_path,
            year=int(year),
            source_filename=source_filename,
            auto_activate=auto_activate,
            uploaded_by=uploaded_by,
        )

    try:
        parsed = parse_survey_excel(excel_path)
    except Exception as exc:
        try:
            mark_batch_failed_by_path(
                str(resolved_db_path),
                excel_path,
                "survey",
                str(exc),
                original_filename=source_filename or os.path.basename(excel_path),
                faculty_id=int(faculty_id),
                year=int(year),
            )
        except Exception:
            pass
        return {"ok": False, "message": f"Anket dosyasi okunamadi: {exc}", "errors": [str(exc)]}

    conn = connect_sqlite(str(resolved_db_path), row_factory=True)
    try:
        ensure_reporting_schema(conn)
        ensure_survey_import_schema(conn)
        cur = conn.cursor()
        try:
            metadata = extract_excel_metadata(excel_path)
        except Exception:
            metadata = {
                "sheet_names": [str(parsed.get("sheet_name") or "Anket")],
                "columns": [],
                "row_count": len(parsed.get("rows") or []),
                "column_count": 0,
            }
        batch = create_import_batch(
            conn,
            import_type="survey",
            original_filename=source_filename or os.path.basename(excel_path),
            file_path=excel_path,
            sheet_names=list(metadata.get("sheet_names") or []),
            columns=list(metadata.get("columns") or []),
            row_count=int(metadata.get("row_count") or len(parsed.get("rows") or [])),
            column_count=int(metadata.get("column_count") or 0),
            faculty_id=int(faculty_id),
            year=int(year),
            uploaded_by=uploaded_by,
            status="uploaded",
        )
        import_batch_id = int(batch["import_batch_id"])
        conn.commit()
        faculty_name = _resolve_faculty_name(cur, int(faculty_id))
        if not faculty_name:
            record_import_issue(
                conn,
                import_batch_id=import_batch_id,
                row_number=0,
                severity="critical",
                issue_type="invalid_scope",
                message="Secili fakulte bulunamadi.",
                suggestion="Import kapsaminda gecerli bir fakulte secin.",
            )
            update_import_status(conn, import_batch_id, "failed", error_message="Secili fakulte bulunamadi.")
            conn.commit()
            return {"ok": False, "message": "Secili fakulte bulunamadi.", "errors": ["Secili fakulte bulunamadi."]}

        # Cok-fakulteli belgeyi secili fakulteye gore filtrele: yalnizca bu fakulteye ait
        # (veya fakulte bilgisi tasimayan) satirlar isleme alinir.
        survey_rows = _filter_rows_for_faculty(list(parsed.get("rows") or []), faculty_name)

        meta = dict(parsed.get("meta") or {})
        declared_total = meta.get("toplam_katilimci")
        validation = validate_survey_rows(
            rows=survey_rows,
            faculty_name=faculty_name,
            year=int(year),
            declared_total_participants=declared_total,
        )
        warnings = list(parsed.get("warnings") or []) + list(validation.get("warnings") or [])
        if not validation.get("ok"):
            for error in list(validation.get("errors") or []):
                record_import_issue(conn, import_batch_id=import_batch_id, row_number=0, message=str(error))
            quality = evaluate_import_quality(conn, import_batch_id)
            update_import_status(
                conn,
                import_batch_id,
                "failed",
                error_message="Anket belgesi dogrulamasi basarisiz.",
                validation_summary={"ok": False, "errors": list(validation.get("errors") or []), "warnings": warnings},
            )
            conn.commit()
            return {
                "ok": False,
                "message": "Anket belgesi dogrulamasi basarisiz.",
                "errors": list(validation.get("errors") or []),
                "warnings": warnings,
                "matched_count": 0,
                "unmatched_count": 0,
                "import_batch_id": import_batch_id,
                "quality_score": quality.quality_score,
                "quality_level": quality.quality_level,
            }

        matched = match_courses(
            conn=conn,
            rows=survey_rows,
            faculty_id=int(faculty_id),
            year=int(year),
        )
        if not matched.get("ok"):
            for unmatched in matched.get("unmatched_rows") or []:
                row_dict = unmatched.as_dict() if hasattr(unmatched, "as_dict") else {}
                record_import_issue(
                    conn,
                    import_batch_id=import_batch_id,
                    row_number=int(row_dict.get("row_no") or 0),
                    issue_type="course_not_matched",
                    severity="error",
                    message=row_dict.get("error_message") or "Ders sistemde bulunamadi.",
                    suggestion="Ders kodu veya adini sistemdeki ders kaydi ile uyumlu hale getirin.",
                )
            quality = evaluate_import_quality(conn, import_batch_id)
            update_import_status(
                conn,
                import_batch_id,
                "pending_review",
                error_message="Belgedeki bazi dersler sistemde bulunamadi.",
                validation_summary={
                    "ok": False,
                    "matched_count": int(matched.get("matched_count") or 0),
                    "unmatched_count": int(matched.get("unmatched_count") or 0),
                },
            )
            conn.commit()
            return {
                "ok": False,
                "message": "Belgedeki bazi dersler sistemde bulunamadi. Veri uygulanmadi.",
                "errors": list(matched.get("errors") or []),
                "warnings": warnings,
                "matched_count": int(matched.get("matched_count") or 0),
                "unmatched_count": int(matched.get("unmatched_count") or 0),
                "matched_rows": [row.as_dict() for row in matched.get("matched_rows") or []],
                "unmatched_rows": [row.as_dict() for row in matched.get("unmatched_rows") or []],
                "import_batch_id": import_batch_id,
                "quality_score": quality.quality_score,
                "quality_level": quality.quality_level,
            }

        matched_rows = list(matched.get("matched_rows") or [])
        declared_total = validation.get("declared_total_participants")
        # Anket cok-secimli olabilir: her katilimci birden fazla ders secebilir. Bu yuzden
        # "toplam_katilimci" bildirilmisse katilimci sayisi odur (tercihlerin toplami DEGIL).
        # Bildirilmemisse tek-secimli kabul edilip tercih toplami katilimci sayilir.
        sum_preferences = compute_total_participants(matched_rows)
        participants = int(declared_total) if declared_total is not None else int(sum_preferences)
        max_pref = max((int(r.tercih_sayisi) for r in matched_rows), default=0)
        # Tutarlilik kontrolu: bir ders, katilimci sayisindan fazla kisi tarafindan secilemez.
        if participants > 0 and max_pref > participants:
            record_import_issue(
                conn,
                import_batch_id=import_batch_id,
                row_number=0,
                issue_type="invalid_numeric_value",
                severity="error",
                message=(
                    "Bir dersin tercih sayisi toplam katilimci sayisindan buyuk olamaz. "
                    f"toplam_katilimci={participants}, en yuksek tercih={max_pref}."
                ),
                suggestion="toplam_katilimci degerini en az en yuksek tercih sayisi kadar girin.",
            )
            quality = evaluate_import_quality(conn, import_batch_id)
            update_import_status(conn, import_batch_id, "pending_review", error_message="Toplam katilimci tutarsizligi.")
            conn.commit()
            return {
                "ok": False,
                "message": "Bir dersin tercih sayisi toplam katilimci sayisindan buyuk olamaz.",
                "errors": [
                    f"toplam_katilimci={participants}, en yuksek tercih={max_pref}."
                ],
                "warnings": warnings,
                "matched_count": int(matched.get("matched_count") or 0),
                "unmatched_count": int(matched.get("unmatched_count") or 0),
                "import_batch_id": import_batch_id,
                "quality_score": quality.quality_score,
                "quality_level": quality.quality_level,
            }

        conn.execute("BEGIN")
        applied = apply_survey_to_criteria(
            conn=conn,
            faculty_id=int(faculty_id),
            year=int(year),
            rows=matched_rows,
            source_filename=source_filename or os.path.basename(excel_path),
            template_version=str(parsed.get("template_version") or SURVEY_TEMPLATE_VERSION),
            notes=meta.get("aciklama"),
            import_batch_id=import_batch_id,
            total_participants_override=participants,
        )
        link_source_import(
            conn,
            import_batch_id=import_batch_id,
            source_table="survey_import",
            source_import_id=int(applied["import_id"]),
            file_hash_sha256=batch.get("file_hash_sha256"),
            file_size=batch.get("file_size"),
        )
        quality = evaluate_import_quality(conn, import_batch_id)
        status = "pending_review" if quality.quality_level == "low" else "validated"
        update_import_status(
            conn,
            import_batch_id,
            status,
            validation_summary={
                "ok": True,
                "matched_count": int(matched.get("matched_count") or 0),
                "unmatched_count": int(matched.get("unmatched_count") or 0),
                "warnings": warnings,
            },
        )
        if auto_activate and quality.quality_level != "low":
            from app.services.import_audit_service import activate_import

            activate_import(conn, import_batch_id, user=uploaded_by)
        try:
            recalculate_import_diff(conn, import_batch_id)
        except Exception:
            pass
        try:
            recalculate_import_impact(conn, import_batch_id)
        except Exception:
            pass
        conn.commit()

        return {
            "ok": True,
            "message": (
                "Anket belgesi basariyla yuklendi. "
                f"Toplam katilimci sayisi otomatik hesaplandi: {applied['total_participants']}."
            ),
            "faculty_id": int(faculty_id),
            "year": int(year),
            "total_participants": int(applied["total_participants"]),
            "matched_count": int(matched.get("matched_count") or 0),
            "unmatched_count": int(matched.get("unmatched_count") or 0),
            "updated_course_count": int(applied.get("updated_rows") or 0),
            "created_course_count": int(applied.get("created_rows") or 0),
            "locked_course_count": int(applied.get("scope_course_count") or 0),
            "replace": dict(applied.get("replace") or {}),
            "warnings": warnings,
            "matched_rows": [row.as_dict() for row in matched_rows],
            "unmatched_rows": [],
            "import_id": int(applied["import_id"]),
            "import_batch_id": import_batch_id,
            "import_status": "active" if auto_activate and quality.quality_level != "low" else status,
            "quality_score": quality.quality_score,
            "quality_level": quality.quality_level,
            "duplicate": bool(batch.get("duplicate")),
            "duplicate_of_import_batch_id": batch.get("duplicate_of_import_batch_id"),
        }
    except Exception as exc:
        conn.rollback()
        try:
            failed_batch_id = locals().get("import_batch_id")
            if failed_batch_id is not None:
                update_import_status(conn, int(failed_batch_id), "failed", error_message=str(exc))
                conn.commit()
        except Exception:
            pass
        return {
            "ok": False,
            "message": f"Anket yukleme hatasi: {exc}",
            "errors": [str(exc)],
            "warnings": list(parsed.get("warnings") or []),
        }
    finally:
        conn.close()


SURVEY_TEMPLATE_COLUMNS = (
    "fakulte_adi",
    "yil",
    "donem",
    "ders_kodu",
    "ders_adi",
    "toplam_katilimci",
    "tercih_sayisi",
)


def _survey_template_row(
    faculty_text: str,
    year_value: int,
    course: dict[str, Any],
) -> dict[str, Any]:
    return {
        "fakulte_adi": faculty_text,
        "yil": year_value,
        "donem": course.get("donem"),
        "ders_kodu": course.get("ders_kodu"),
        "ders_adi": course.get("ders_adi"),
        "toplam_katilimci": None,
        "tercih_sayisi": None,
    }


def write_survey_template_excel(
    target_path: str,
    faculty_name: str | None = None,
    year: int | None = None,
    db_path: str | None = None,
    faculty_id: int | None = None,
) -> str:
    """Anket/tercih sablonu uretir.

    - faculty_id verilirse: yalnizca o fakultenin havuzdaki secmeli dersleri yazilir.
    - faculty_id None ve db_path verilirse ('Tumu'): tum fakultelerin dersleri yazilir.
    - db_path yoksa: ornek satirlardan olusan bos sablon yazilir.

    Kolon duzeni gercek anket veri seti ile birebir ayni:
    fakulte_adi | yil | donem | ders_kodu | ders_adi | toplam_katilimci | tercih_sayisi
    """
    year_value = int(year or datetime.now().year)
    survey_rows: list[dict[str, Any]] = []
    multi_faculty = False

    if db_path and faculty_id is not None and year is not None:
        context = load_survey_template_context(
            db_path=db_path,
            faculty_id=int(faculty_id),
            year=int(year),
        )
        faculty_name = str(context.get("faculty_name") or faculty_name or "")
        template_courses = list(context.get("courses") or [])
        if not template_courses:
            raise ValueError("Secili fakulte ve yil icin havuzda uygun secmeli ders bulunamadi.")
        survey_rows = [_survey_template_row(faculty_name, year_value, c) for c in template_courses]
    elif db_path and faculty_id is None and year is not None:
        # "Tumu": tum fakulteleri tek dosyada birlestir.
        contexts = load_all_faculties_template_context(db_path=db_path, year=int(year))
        if not contexts:
            raise ValueError("Secili yil icin hicbir fakultenin havuzunda uygun secmeli ders bulunamadi.")
        multi_faculty = True
        for ctx in contexts:
            fac = str(ctx.get("faculty_name") or "")
            for course in list(ctx.get("courses") or []):
                survey_rows.append(_survey_template_row(fac, year_value, course))
    else:
        faculty_text = faculty_name or "Ornek Fakultesi"
        survey_rows = [
            _survey_template_row(faculty_text, year_value, {"ders_kodu": "SEC101", "ders_adi": "Ornek Secmeli Ders 1", "donem": "Güz"}),
            _survey_template_row(faculty_text, year_value, {"ders_kodu": "SEC102", "ders_adi": "Ornek Secmeli Ders 2", "donem": "Bahar"}),
        ]

    survey_df = pd.DataFrame(survey_rows, columns=list(SURVEY_TEMPLATE_COLUMNS))

    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        survey_df.to_excel(writer, sheet_name=SURVEY_TEMPLATE_SHEET_NAME, index=False)
        worksheet = writer.sheets[SURVEY_TEMPLATE_SHEET_NAME]
        header_columns = list(survey_df.columns)
        pref_col_idx = header_columns.index("tercih_sayisi") + 1
        ders_adi_col_idx = header_columns.index("ders_adi") + 1
        first_data_row = 2
        last_data_row = len(survey_rows) + 1
        pref_letter = _excel_column_letter(pref_col_idx)

        # Cok-fakulteli sablonda tek bir genel TOPLAM anlamsiz olur; yalnizca tek
        # fakulte sablonunda toplam satiri eklenir.
        if not multi_faculty and survey_rows:
            total_row = last_data_row + 1
            worksheet.cell(row=total_row, column=ders_adi_col_idx, value="TOPLAM")
            worksheet.cell(
                row=total_row,
                column=pref_col_idx,
                value=f"=SUM({pref_letter}{first_data_row}:{pref_letter}{last_data_row})",
            )
            worksheet.cell(row=total_row, column=ders_adi_col_idx).font = Font(bold=True)
            worksheet.cell(row=total_row, column=pref_col_idx).font = Font(bold=True)
    return target_path
